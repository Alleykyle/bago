from django.shortcuts import render, redirect, get_object_or_404
from .models import Employee
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate, login, logout
import json
from django.contrib.auth.models import User
from .models import UserProfile
from django.core.exceptions import ValidationError
from django.contrib.auth.password_validation import validate_password
from .decorators import role_required
from django.db.models import Q, Count, Avg, Sum
from django.db import transaction, connection
from django.core.paginator import Paginator
from django.core.cache import cache
from django.http import HttpResponse, HttpResponseForbidden
import csv
import openpyxl
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Q, Count, Avg, Sum  
from .models import EligibilityRequest, Barangay, Requirement, RequirementSubmission, RequirementAttachment, Notification, Announcement, FileCategory, MonitoringFile
from django.views.decorators.http import require_POST
import os
from datetime import date
from .models import CategorizedFile
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from datetime import timedelta
import traceback
from PIL import Image
import pytesseract, PyPDF2



try:
    from .models import AuditLog
except ImportError:
    class AuditLog:
        @staticmethod
        def objects():
            return None
        
        @staticmethod
        def create(**kwargs):
            pass


def get_client_ip(request):
    """Helper function to get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def landing_page(request):
    return render(request, 'landing.html')


def logout_view(request):
    if request.user.is_authenticated:
        try:
            AuditLog.objects.create(
                user=request.user,
                action='LOGOUT',
                ip_address=get_client_ip(request),
                description=f"User {request.user.username} logged out"
            )
        except:
            pass  # Skip audit logging if AuditLog doesn't exist
    
    logout(request)
    messages.success(request, "You have successfully logged out.")
    return redirect('landing_page')


def login_page(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        ip_address = get_client_ip(request)
        user_agent = request.META.get('HTTP_USER_AGENT', 'Unknown')

        if user is not None:
            login(request, user)
            profile, created = UserProfile.objects.get_or_create(user=user)
            profile.update_login_info(ip_address)

            AuditLog.objects.create(
                user=user,
                action='LOGIN',
                ip_address=ip_address,
                user_agent=user_agent,
                description=f"{user.username} logged in as {profile.role}",
            )

            # 🧠 Debug
            print("🔍 ROLE DETECTED:", profile.role)
            print("🔍 REDIRECTING TO:", profile.get_redirect_url())

            redirect_url = profile.get_redirect_url()
            messages.success(request, f"Welcome back, {user.username}!")
            return redirect(redirect_url)

        else:
            messages.error(request, "Invalid username or password.")

    return render(request, 'login_page.html')


@login_required
def logout_view(request):
    """Handle user logout with audit log"""
    user = request.user
    ip_address = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', 'Unknown')

    # Log the logout action
    AuditLog.objects.create(
        user=user,
        action='LOGOUT',
        ip_address=ip_address,
        user_agent=user_agent,
        description=f"{user.username} logged out",
    )

    logout(request)
    messages.info(request, "You have been logged out successfully.")
    return redirect('login')  # change to your login page URL name


def landing_menu(request):
    return render(request, 'landing_menu.html')



def dashboard(request):
    """
    Safe Analytics Dashboard View - No attribute errors
    """
    try:
        # Basic Employee Statistics - These should always work
        total_employees = Employee.objects.count()
        active_employees = Employee.objects.filter(status='active').count()
        
        # Simple stats that don't depend on unknown fields
        context = {
            'total_employees': total_employees,
            'active_employees': active_employees,
            'new_employees_30d': 0,  # We'll calculate this safely later
            'dept_stats': [],  # Empty for now
            'task_stats': [],  # Empty for now
            'user_activity': [],  # Empty for now
            'recent_activities': [],  # Empty for now
            'table_stats': [('Employees', total_employees)],
            'db_size': None,
        }
        
        print(f"Dashboard stats: Total={total_employees}, Active={active_employees}")
        return render(request, 'dashboard.html', context)
        
    except Exception as e:
        print(f"Dashboard error: {e}")
        # Return safe empty context
        context = {
            'total_employees': 0,
            'active_employees': 0,
            'new_employees_30d': 0,
            'dept_stats': [],
            'task_stats': [],
            'user_activity': [],
            'recent_activities': [],
            'table_stats': [],
            'db_size': None,
        }
        return render(request, 'dashboard.html', context)


@login_required
def refresh_analytics(request):
    """
    Safe AJAX endpoint to refresh analytics data
    """
    if request.method == 'GET':
        try:
            total_employees = Employee.objects.count()
            active_employees = Employee.objects.filter(status='active').count()
            
            return JsonResponse({
                'success': True,
                'data': {
                    'total_employees': total_employees,
                    'active_employees': active_employees,
                    'timestamp': timezone.now().strftime('%B %d, %H:%M')
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


# DEBUG VIEW - Add this temporarily to see your model structure
@login_required 
def debug_employee(request):
    """
    Debug view to inspect Employee model
    """
    try:
        # Get first employee
        emp = Employee.objects.first()
        
        # Get all field names
        field_names = [f.name for f in Employee._meta.fields]
        
        # Get sample data safely
        sample_data = {}
        if emp:
            for field_name in field_names:
                try:
                    value = getattr(emp, field_name, 'N/A')
                    sample_data[field_name] = str(value)[:100]  # Limit length
                except:
                    sample_data[field_name] = 'Error reading'
        
        return JsonResponse({
            'total_employees': Employee.objects.count(),
            'field_names': field_names,
            'sample_data': sample_data
        }, indent=2)
        
    except Exception as e:
        return JsonResponse({'error': str(e)})



def application_request(request):
    return render(request, 'application_request.html')


def history(request):
    return render(request, 'history.html')

@login_required
def history_api(request):
    """AJAX endpoint for history page real-time updates"""
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)
    
    try:
        # Get the last activity ID from the request to check for new activities
        last_activity_id = request.GET.get('last_id', 0)
        
        # Check if there are new activities since the last check
        new_activities_count = AuditLog.objects.filter(id__gt=last_activity_id).count()
        
        # Get activity statistics
        today = timezone.now().date()
        stats = {
            'total_activities': AuditLog.objects.count(),
            'today_activities': AuditLog.objects.filter(timestamp__date=today).count(),
            'unique_users': AuditLog.objects.exclude(user__isnull=True).values('user').distinct().count(),
            'action_types': AuditLog.objects.values('action').distinct().count(),
            'has_new_activities': new_activities_count > 0,
            'new_activities_count': new_activities_count
        }
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
@login_required
def export_history(request):
    """Export history data as CSV or Excel"""
    try:
        format_type = request.GET.get('format', 'csv')
        
        # Apply same filters as history view
        activities = AuditLog.objects.select_related('user', 'content_type').order_by('-timestamp')
        
        # Apply filters
        action_filter = request.GET.get('action', '')
        user_filter = request.GET.get('user', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        search_query = request.GET.get('search', '')
        
        if action_filter:
            activities = activities.filter(action=action_filter)
        
        if user_filter:
            activities = activities.filter(user__username__icontains=user_filter)
        
        if date_from:
            activities = activities.filter(timestamp__date__gte=date_from)
        
        if date_to:
            activities = activities.filter(timestamp__date__lte=date_to)
        
        if search_query:
            activities = activities.filter(
                Q(description__icontains=search_query) |
                Q(user__username__icontains=search_query) |
                Q(action__icontains=search_query)
            )
        
        # Limit export to prevent performance issues
        activities = activities[:5000]  # Max 5000 records
        
        if format_type == 'excel':
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="activity_history_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Activity History'
            
            # Headers
            headers = [
                'ID', 'User', 'Action', 'Description', 'Timestamp', 
                'IP Address', 'Content Type', 'Object ID', 'Old Values', 'New Values'
            ]
            
            # Style headers
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center')
            
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Data rows
            for row, activity in enumerate(activities, 2):
                worksheet.cell(row=row, column=1, value=activity.id)
                worksheet.cell(row=row, column=2, value=activity.user.username if activity.user else 'System')
                worksheet.cell(row=row, column=3, value=activity.action)
                worksheet.cell(row=row, column=4, value=activity.description)
                worksheet.cell(row=row, column=5, value=activity.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
                worksheet.cell(row=row, column=6, value=activity.ip_address or '')
                worksheet.cell(row=row, column=7, value=str(activity.content_type) if activity.content_type else '')
                worksheet.cell(row=row, column=8, value=activity.object_id or '')
                worksheet.cell(row=row, column=9, value=json.dumps(activity.old_values) if activity.old_values else '')
                worksheet.cell(row=row, column=10, value=json.dumps(activity.new_values) if activity.new_values else '')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(response)
            
        else:  # CSV format
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="activity_history_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
            
            writer = csv.writer(response)
            writer.writerow([
                'ID', 'User', 'Action', 'Description', 'Timestamp', 
                'IP Address', 'Content Type', 'Object ID', 'Old Values', 'New Values'
            ])
            
            for activity in activities:
                writer.writerow([
                    activity.id,
                    activity.user.username if activity.user else 'System',
                    activity.action,
                    activity.description,
                    activity.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    activity.ip_address or '',
                    str(activity.content_type) if activity.content_type else '',
                    activity.object_id or '',
                    json.dumps(activity.old_values) if activity.old_values else '',
                    json.dumps(activity.new_values) if activity.new_values else ''
                ])
        
        # Log the export activity
        try:
            AuditLog.objects.create(
                user=request.user,
                action='EXPORT',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                description=f"Exported activity history as {format_type.upper()}",
                new_values={
                    'export_format': format_type,
                    'record_count': activities.count(),
                    'filters_applied': {
                        'action': action_filter,
                        'user': user_filter,
                        'date_from': date_from,
                        'date_to': date_to,
                        'search': search_query
                    }
                }
            )
        except:
            pass
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Export failed: {str(e)}'
        }, status=500)

@login_required
def bulk_history_operations(request):
    """Bulk operations on history records (for admins)"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    try:
        data = json.loads(request.body)
        action = data.get('action')
        activity_ids = data.get('activity_ids', [])
        
        if not activity_ids:
            return JsonResponse({'success': False, 'error': 'No activities selected'})
        
        activities = AuditLog.objects.filter(id__in=activity_ids)
        count = activities.count()
        
        if action == 'delete':
            # Store info about deleted activities for logging
            deleted_info = list(activities.values('id', 'action', 'description', 'user__username'))
            activities.delete()
            
            # Log the bulk deletion
            AuditLog.objects.create(
                user=request.user,
                action='BULK_DELETE',
                ip_address=get_client_ip(request),
                description=f"Bulk deleted {count} activity records",
                old_values={'deleted_activities': deleted_info}
            )
            
            return JsonResponse({
                'success': True,
                'message': f'{count} activities deleted successfully'
            })
        
        elif action == 'archive':
            # Mark as archived (you could add an 'archived' field to AuditLog)
            # For now, just add a note to the description
            activities.update(description=F('description') + ' [ARCHIVED]')
            
            AuditLog.objects.create(
                user=request.user,
                action='BULK_ARCHIVE',
                ip_address=get_client_ip(request),
                description=f"Bulk archived {count} activity records",
                new_values={'archived_count': count, 'activity_ids': activity_ids}
            )
            
            return JsonResponse({
                'success': True,
                'message': f'{count} activities archived successfully'
            })
        
        else:
            return JsonResponse({'success': False, 'error': 'Invalid action'})
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Bulk operation failed: {str(e)}'
        }, status=500)
    
@login_required
def activity_stats(request):
    """Get activity statistics for the dashboard"""
    try:
        today = timezone.now().date()
        yesterday = today - timedelta(days=1)
        last_7_days = today - timedelta(days=7)
        last_30_days = today - timedelta(days=30)
        
        stats = {
            'total_activities': AuditLog.objects.count(),
            'today_activities': AuditLog.objects.filter(timestamp__date=today).count(),
            'yesterday_activities': AuditLog.objects.filter(timestamp__date=yesterday).count(),
            'week_activities': AuditLog.objects.filter(timestamp__date__gte=last_7_days).count(),
            'month_activities': AuditLog.objects.filter(timestamp__date__gte=last_30_days).count(),
            'unique_users': AuditLog.objects.exclude(user__isnull=True).values('user').distinct().count(),
            'action_types': AuditLog.objects.values('action').distinct().count(),
        }
        
        # Activity breakdown by action type
        action_breakdown = list(
            AuditLog.objects.values('action')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )
        
        # Daily activity trend for the last 7 days
        daily_trend = []
        for i in range(7):
            date = today - timedelta(days=i)
            count = AuditLog.objects.filter(timestamp__date=date).count()
            daily_trend.append({
                'date': date.strftime('%Y-%m-%d'),
                'count': count
            })
        
        # Most active users
        active_users = list(
            AuditLog.objects.exclude(user__isnull=True)
            .values('user__username')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )
        
        stats.update({
            'action_breakdown': action_breakdown,
            'daily_trend': daily_trend,
            'active_users': active_users
        })
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)




def employees_profile(request):
    # Handle POST - Add new employee with validation
    if request.method == 'POST':
        try:
            with transaction.atomic():  # Ensure data consistency
                name = request.POST.get('name', '').strip()
                id_no = request.POST.get('id_no', '').strip()
                task = request.POST.get('task', '').strip()
                email = request.POST.get('email', '').strip()
                phone = request.POST.get('phone', '').strip()
                department = request.POST.get('department', '')
                position = request.POST.get('position', '').strip()
                hire_date = request.POST.get('hire_date')
                supervisor_id = request.POST.get('supervisor')
                
                # Validation
                if not all([name, id_no, task]):
                    return JsonResponse({
                        'success': False, 
                        'error': 'Name, ID number, and task are required'
                    }, status=400)
                
                # Check for duplicate ID
                if Employee.objects.filter(id_no=id_no).exists():
                    return JsonResponse({
                        'success': False, 
                        'error': 'Employee ID already exists'
                    }, status=400)
                
                # Create employee
                employee_data = {
                    'name': name,
                    'id_no': id_no,
                    'task': task,
                    'department': department if department else None,
                    'position': position,
                }
                
                if email:
                    employee_data['email'] = email
                if phone:
                    employee_data['phone'] = phone
                if hire_date:
                    employee_data['hire_date'] = hire_date
                if supervisor_id:
                    employee_data['supervisor_id'] = supervisor_id
                
                employee = Employee.objects.create(**employee_data)
                
                messages.success(request, f'Employee {name} added successfully!')
                
                # Clear cache
                cache.delete('employee_stats')
                
                return redirect('employees_profile')
                
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': f'Error creating employee: {str(e)}'
            }, status=400)
    
    # Handle GET - Display employees with advanced filtering and pagination
    search_query = request.GET.get('search', '').strip()
    department_filter = request.GET.get('department', '')
    status_filter = request.GET.get('status', '')
    sort_by = request.GET.get('sort', 'name')
    
    # Base queryset with optimizations
    employees = Employee.objects.select_related('supervisor').prefetch_related('subordinates')
    
    # Apply filters
    if search_query:
        employees = employees.filter(
            Q(name__icontains=search_query) |
            Q(id_no__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(task__icontains=search_query) |
            Q(position__icontains=search_query)
        )
    
    if department_filter:
        employees = employees.filter(department=department_filter)
    
    if status_filter:
        employees = employees.filter(status=status_filter)
    
    # Sorting
    valid_sort_fields = ['name', '-name', 'id_no', '-id_no', 'created_at', '-created_at', 'department']
    if sort_by in valid_sort_fields:
        employees = employees.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(employees, 10)  # Show 10 employees per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get statistics (cached for performance)
    stats = cache.get('employee_stats')
    if not stats:
        try:
            stats = Employee.get_statistics()
        except AttributeError:
            # If get_statistics method doesn't exist, create basic stats
            stats = {
                'total': Employee.objects.count(),
                'active': Employee.objects.filter(status='active').count() if hasattr(Employee, 'status') else 0,
                'departments': Employee.objects.values('department').distinct().count()
            }
        cache.set('employee_stats', stats, 300)  # Cache for 5 minutes
    
    # Get choices for dropdowns
    supervisors = Employee.objects.filter(status='active') if hasattr(Employee, 'status') else Employee.objects.all()
    
    # Get department and status choices if they exist
    department_choices = getattr(Employee, 'DEPARTMENT_CHOICES', [])
    status_choices = getattr(Employee, 'STATUS_CHOICES', [])
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'department_filter': department_filter,
        'status_filter': status_filter,
        'sort_by': sort_by,
        'stats': stats,
        'supervisors': supervisors,
        'department_choices': department_choices,
        'status_choices': status_choices,
    }
    
    return render(request, 'employees_profile.html', context)


@require_http_methods(["POST"])
def edit_employee(request, employee_id):
    """Handle AJAX edit employee requests"""
    try:
        employee = get_object_or_404(Employee, pk=employee_id)
        
        # Parse JSON data from request body
        data = json.loads(request.body)
        
        name = data.get('name', '').strip()
        id_no = data.get('id_no', '').strip()
        task = data.get('task', '').strip()
        
        # Validate required fields
        if not name:
            return JsonResponse({
                'success': False, 
                'error': 'Name is required'
            }, status=400)
        
        if not id_no:
            return JsonResponse({
                'success': False, 
                'error': 'ID number is required'
            }, status=400)
        
        if not task:
            return JsonResponse({
                'success': False, 
                'error': 'Task assignment is required'
            }, status=400)
        
        # Check if ID number is already taken by another employee
        existing_employee = Employee.objects.filter(id_no=id_no).exclude(pk=employee_id).first()
        if existing_employee:
            return JsonResponse({
                'success': False, 
                'error': 'This ID number is already assigned to another employee'
            }, status=400)
        
        # Update employee
        employee.name = name
        employee.id_no = id_no
        employee.task = task
        employee.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Employee updated successfully'
        })
        
    except Employee.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'Employee not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred: {str(e)}'
        }, status=500)


@require_http_methods(["DELETE"])
def delete_employee(request, employee_id):
    try:
        employee = Employee.objects.get(id=employee_id)
        employee_name = employee.name  # Store name for response
        employee.delete()
        return JsonResponse({
            'success': True,
            'message': f'Employee {employee_name} deleted successfully'
        })
    except Employee.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'Employee not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred: {str(e)}'
        }, status=500)


def folder(request):
    return render(request, 'folder.html')


def settings(request):
    return render(request, 'settings.html')



def civil_service_certification(request):
    """Render the public certification form"""
    return render(request, 'civil_service_certification.html')


def application_letter(request):
    return render(request, 'application_letter.html')


def monitoring_filess(request):
    return render(request, 'monitoring_files.html')


def certification_filess(request):
    return render(request, 'certification_files.html')



def signup_page(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        role = request.POST.get('role')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        if not all([username, email, role, password1, password2]):
            messages.error(request, 'Please fill in all fields.')
            return render(request, 'signup_page.html')

        if password1 != password2:
            messages.error(request, 'Passwords do not match.')
            return render(request, 'signup_page.html')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return render(request, 'signup_page.html')

        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email already registered.')
            return render(request, 'signup_page.html')

        try:
            validate_password(password1)
        except ValidationError as e:
            for error in e:
                messages.error(request, error)
            return render(request, 'signup_page.html')

        # Create the user
        user = User.objects.create_user(username=username, email=email, password=password1)

        # Create or get profile with normalized role (strip spaces)
        UserProfile.objects.get_or_create(user=user, defaults={'role': role.strip()})

        messages.success(request, 'Account created successfully. Please log in.')
        return redirect('login_page')

    return render(request, 'signup_page.html')


@login_required
def export_employees(request):
    """Export employees data"""
    format_type = request.GET.get('format', 'csv')
    
    employees = Employee.objects.select_related('supervisor').all()
    
    if format_type == 'excel':
        # Create Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="employees_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Employees'
        
        # Headers
        headers = ['ID', 'Name', 'Employee ID', 'Email', 'Department', 'Position', 'Status', 'Task', 'Supervisor', 'Hire Date', 'Created At']
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Data
        for row, employee in enumerate(employees, 2):
            worksheet.cell(row=row, column=1, value=employee.id)
            worksheet.cell(row=row, column=2, value=employee.name)
            worksheet.cell(row=row, column=3, value=employee.id_no)
            worksheet.cell(row=row, column=4, value=getattr(employee, 'email', '') or '')
            
            # Handle department display
            if hasattr(employee, 'get_department_display'):
                dept = employee.get_department_display() or ''
            else:
                dept = getattr(employee, 'department', '') or ''
            worksheet.cell(row=row, column=5, value=dept)
            
            worksheet.cell(row=row, column=6, value=getattr(employee, 'position', '') or '')
            
            # Handle status display
            if hasattr(employee, 'get_status_display'):
                status = employee.get_status_display()
            else:
                status = getattr(employee, 'status', '') or ''
            worksheet.cell(row=row, column=7, value=status)
            
            worksheet.cell(row=row, column=8, value=getattr(employee, 'task', '') or '')
            worksheet.cell(row=row, column=9, value=employee.supervisor.name if employee.supervisor else '')
            worksheet.cell(row=row, column=10, value=getattr(employee, 'hire_date', ''))
            worksheet.cell(row=row, column=11, value=getattr(employee, 'created_at', ''))
        
        workbook.save(response)
        
    else:  # CSV format
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="employees_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Name', 'Employee ID', 'Email', 'Department', 'Position', 'Status', 'Task', 'Supervisor', 'Hire Date', 'Created At'])
        
        for employee in employees:
            # Handle department display
            if hasattr(employee, 'get_department_display'):
                dept = employee.get_department_display() or ''
            else:
                dept = getattr(employee, 'department', '') or ''
            
            # Handle status display
            if hasattr(employee, 'get_status_display'):
                status = employee.get_status_display()
            else:
                status = getattr(employee, 'status', '') or ''
            
            writer.writerow([
                employee.id,
                employee.name,
                employee.id_no,
                getattr(employee, 'email', '') or '',
                dept,
                getattr(employee, 'position', '') or '',
                status,
                getattr(employee, 'task', '') or '',
                employee.supervisor.name if employee.supervisor else '',
                getattr(employee, 'hire_date', ''),
                getattr(employee, 'created_at', '')
            ])
    
    # Log export action
    try:
        AuditLog.objects.create(
            user=request.user,
            action='CREATE',
            description=f"Exported employees data as {format_type.upper()}"
        )
    except:
        pass  # Skip audit logging if AuditLog doesn't exist
    
    return response


@login_required
@role_required('dilg staff')
def analytics_dashboard(request):
    """Advanced analytics dashboard"""
    
    # Time-based statistics
    today = timezone.now().date()
    last_30_days = today - timedelta(days=30)
    last_7_days = today - timedelta(days=7)
    
    # Employee metrics
    total_employees = Employee.objects.count()
    print(f"DEBUG: Total employees found: {total_employees}")
    
    # Fix for new employees - check if created_at field exists and has data
    if hasattr(Employee, 'created_at'):
        new_employees_30d = Employee.objects.filter(created_at__date__gte=last_30_days).count()
    else:
        new_employees_30d = 0
    
    # Fix for active employees - count all OR only active status
    if hasattr(Employee, 'status'):
        # Count employees with active status OR null/empty status (treat as active)
        active_employees = Employee.objects.filter(
            Q(status='active') | Q(status__isnull=True) | Q(status='')
        ).count()
        print(f"DEBUG: Active employees: {active_employees}")
        
        # If still 0, just count all employees as active
        if active_employees == 0:
            active_employees = total_employees
    else:
        active_employees = total_employees
    
    # Department statistics - with better error handling
    dept_stats = []
    try:
        dept_stats = list(Employee.objects.values('department')
                         .annotate(count=Count('id'))
                         .order_by('-count'))
        print(f"DEBUG: Department stats: {dept_stats}")
        
        # Clean up None departments
        for stat in dept_stats:
            if not stat['department']:
                stat['department'] = 'Unassigned'
                
    except Exception as e:
        print(f"DEBUG: Error getting dept stats: {e}")
        pass
    
    # Task distribution - with better error handling
    task_stats = []
    try:
        task_stats = list(Employee.objects.exclude(task__isnull=True)
                         .exclude(task='')
                         .values('task')
                         .annotate(count=Count('id'))
                         .order_by('-count')[:10])
        print(f"DEBUG: Task stats: {task_stats}")
    except Exception as e:
        print(f"DEBUG: Error getting task stats: {e}")
        pass
    
    # Recent activities from audit log
    recent_activities = []
    user_activity = []
    try:
        recent_activities = AuditLog.objects.select_related('user')\
                                          .filter(timestamp__date__gte=last_7_days)\
                                          .order_by('-timestamp')[:20]
        
        # User activity stats
        user_activity = AuditLog.objects.filter(action='LOGIN', timestamp__date__gte=last_30_days)\
                                       .values('user__username')\
                                       .annotate(login_count=Count('id'))\
                                       .order_by('-login_count')[:10]
    except Exception as e:
        print(f"DEBUG: Error getting audit logs: {e}")
        pass
    
    # System health metrics
    db_size = 0
    table_stats = []
    try:
        with connection.cursor() as cursor:
            # Database size (SQLite specific)
            cursor.execute("SELECT page_count * page_size as size FROM pragma_page_count(), pragma_page_size();")
            result = cursor.fetchone()
            if result:
                db_size = result[0]
            
            # Basic table stats
            table_stats = [
                ('Employees', Employee.objects.count()),
                ('Users', User.objects.count()),
            ]
    except Exception as e:
        print(f"DEBUG: Error getting system stats: {e}")
        pass
    
    context = {
        'total_employees': total_employees,
        'new_employees_30d': new_employees_30d,
        'active_employees': active_employees,
        'dept_stats': dept_stats,
        'task_stats': task_stats,
        'recent_activities': recent_activities,
        'user_activity': user_activity,
        'db_size': db_size,
        'table_stats': table_stats,
    }
    
    print(f"DEBUG: Final context: {context}")
    return render(request, 'analytics_dashboard.html', context)


# Advanced search API
@login_required
def employee_search_api(request):
    """AJAX endpoint for advanced employee search"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'employees': []})
    
    # Complex search across multiple fields
    try:
        employees = Employee.objects.filter(
            Q(name__icontains=query) |
            Q(id_no__icontains=query) |
            Q(email__icontains=query) |
            Q(position__icontains=query) |
            Q(department__icontains=query)
        ).values('id', 'name', 'id_no', 'email', 'department', 'status')[:10]
    except:
        employees = Employee.objects.filter(
            Q(name__icontains=query) |
            Q(id_no__icontains=query)
        ).values('id', 'name', 'id_no')[:10]
    
    return JsonResponse({
        'employees': list(employees)
    })


# Bulk operations
@login_required
@require_http_methods(["POST"])
def bulk_employee_operations(request):
    """Handle bulk operations on employees"""
    try:
        data = json.loads(request.body)
        action = data.get('action')
        employee_ids = data.get('employee_ids', [])
        
        if not employee_ids:
            return JsonResponse({'success': False, 'error': 'No employees selected'})
        
        with transaction.atomic():
            employees = Employee.objects.filter(id__in=employee_ids)
            
            if action == 'delete':
                count = employees.count()
                employees.delete()
                message = f'{count} employees deleted successfully'
                
            elif action == 'activate':
                if hasattr(Employee, 'status'):
                    count = employees.update(status='active')
                    message = f'{count} employees activated'
                else:
                    return JsonResponse({'success': False, 'error': 'Status field not available'})
                
            elif action == 'deactivate':
                if hasattr(Employee, 'status'):
                    count = employees.update(status='inactive')
                    message = f'{count} employees deactivated'
                else:
                    return JsonResponse({'success': False, 'error': 'Status field not available'})
                
            elif action == 'update_department':
                department = data.get('department')
                if not department:
                    return JsonResponse({'success': False, 'error': 'Department required'})
                if hasattr(Employee, 'department'):
                    count = employees.update(department=department)
                    message = f'{count} employees moved to {department}'
                else:
                    return JsonResponse({'success': False, 'error': 'Department field not available'})
                
            else:
                return JsonResponse({'success': False, 'error': 'Invalid action'})
            
            # Log bulk operation
            try:
                AuditLog.objects.create(
                    user=request.user,
                    action='UPDATE',
                    description=f"Bulk operation: {action} on {len(employee_ids)} employees"
                )
            except:
                pass  # Skip audit logging if AuditLog doesn't exist
            
            # Clear cache
            cache.delete('employee_stats')
            
            return JsonResponse({'success': True, 'message': message})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    

from PIL import Image, ImageOps
import io
from django.core.files.uploadedfile import InMemoryUploadedFile

def process_signature_image(uploaded_file):
    """
    Process signature image to ensure it has white background
    and black signature (fixes black signature display issue)
    """
    try:
        print(f"🖊️ Processing signature: {uploaded_file.name}")
        print(f"   Original size: {uploaded_file.size} bytes")
        
        # Store original filename
        original_name = uploaded_file.name
        
        # Read the uploaded file
        uploaded_file.seek(0)
        image = Image.open(uploaded_file)
        
        print(f"   Image mode: {image.mode}")
        print(f"   Image size: {image.size}")
        
        # Convert to RGBA first if needed
        if image.mode != 'RGBA':
            image = image.convert('RGBA')
        
        # Create a white background
        white_bg = Image.new('RGB', image.size, (255, 255, 255))
        
        # Paste the signature onto white background
        # This converts transparent areas to white
        white_bg.paste(image, mask=image.split()[-1])  # Use alpha channel as mask
        
        # Save to BytesIO
        output = io.BytesIO()
        white_bg.save(output, format='PNG', quality=95)
        output.seek(0)
        
        processed_size = output.getbuffer().nbytes
        print(f"   ✓ Processed size: {processed_size} bytes")
        
        # Create new InMemoryUploadedFile with ORIGINAL FILENAME
        processed_file = InMemoryUploadedFile(
            output,
            'ImageField',
            original_name,  # ✅ Keep the original filename
            'image/png',
            processed_size,
            None
        )
        
        return processed_file
        
    except Exception as e:
        print(f"   ⚠️ Error processing signature: {e}")
        import traceback
        print(traceback.format_exc())
        # Return original file if processing fails
        uploaded_file.seek(0)
        return uploaded_file



@require_http_methods(["POST"])
def submit_eligibility_request(request):
    """
    Handle form submission with SMART DOCUMENT CATEGORIZATION
    Files are automatically sorted based on content analysis
    """
    try:
        print("\n" + "="*80)
        print("🆕 NEW ELIGIBILITY REQUEST SUBMISSION")
        print("="*80)
        
        # Extract and validate form data
        last_name = request.POST.get('last_name', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        middle_initial = request.POST.get('middle_initial', '').strip()
        barangay = request.POST.get('barangay', '').strip()
        email = request.POST.get('email', '').strip()
        position_type = request.POST.get('position_type', '').strip()
        certifier = request.POST.get('certifier', '').strip()
        
        print(f"👤 Applicant: {first_name} {last_name}")
        print(f"📧 Email: {email}")
        print(f"🏢 Barangay: {barangay}")
        print(f"📋 Position Type: {position_type}")
        
        # Validation
        if not all([last_name, first_name, email, barangay, position_type, certifier]):
            return JsonResponse({
                'success': False,
                'error': 'All required fields must be filled'
            }, status=400)
        
        # Validate certifier choice
        valid_certifiers = [choice[0] for choice in EligibilityRequest.CERTIFIER_CHOICES]
        if certifier not in valid_certifiers:
            return JsonResponse({
                'success': False,
                'error': f'Invalid certifier selection'
            }, status=400)
        
        # Get uploaded files
        id_front = request.FILES.get('id_front')
        id_back = request.FILES.get('id_back')
        signature = request.FILES.get('signature')
        
        if not all([id_front, id_back, signature]):
            return JsonResponse({
                'success': False,
                'error': 'All files (ID front, ID back, signature) are required'
            }, status=400)
        
        print(f"📎 Files received: ID Front ({id_front.size} bytes), ID Back ({id_back.size} bytes), Signature ({signature.size} bytes)")
        
        # Create eligibility request
        eligibility_request = EligibilityRequest.objects.create(
            first_name=first_name,
            last_name=last_name,
            middle_initial=middle_initial if middle_initial else None,
            barangay=barangay,
            email=email,
            position_type=position_type,
            certifier=certifier,
            status='pending',
            date_submitted=timezone.now()
        )
        
        # Save position-specific data
        if position_type == 'appointive':
            eligibility_request.appointing_authority = request.POST.get('appointing_authority', '')
            eligibility_request.appointment_from = request.POST.get('appointment_from')
            eligibility_request.appointment_to = request.POST.get('appointment_to')
            eligibility_request.years_in_service = request.POST.get('years_in_service')
            eligibility_request.appointing_punong_barangay = request.POST.get('appointing_punong_barangay', '')
            eligibility_request.pb_date_elected = request.POST.get('pb_date_elected')
            eligibility_request.pb_years_service = request.POST.get('pb_years_service')
        elif position_type == 'elective':
            eligibility_request.position_held = request.POST.get('position_held', '')
            eligibility_request.election_from = request.POST.get('election_from')
            eligibility_request.election_to = request.POST.get('election_to')
            eligibility_request.term_office = request.POST.get('term_office', '')
            eligibility_request.completed_term = request.POST.get('completed_term', '')
            eligibility_request.incomplete_reason = request.POST.get('incomplete_reason', '')
            eligibility_request.days_not_served = int(request.POST.get('days_not_served', 0))
        
        eligibility_request.save()
        
        print(f"✅ Created EligibilityRequest ID: {eligibility_request.id}")
        
        # 🔥 SMART CATEGORIZATION - Process each file
        files_processed = []
        
        # Process ID Front
        print(f"\n📄 Processing ID Front...")
        id_front_category = smart_categorize_file(id_front, 'id_front')
        id_front_path = save_categorized_eligibility_file(
            file=id_front,
            category=id_front_category,
            user_name=f"{first_name}_{last_name}",
            file_type='id_front',
            request_id=eligibility_request.id
        )
        files_processed.append({
            'name': 'ID Front',
            'category': id_front_category,
            'path': id_front_path
        })
        
        # Process ID Back
        print(f"\n📄 Processing ID Back...")
        id_back_category = smart_categorize_file(id_back, 'id_back')
        id_back_path = save_categorized_eligibility_file(
            file=id_back,
            category=id_back_category,
            user_name=f"{first_name}_{last_name}",
            file_type='id_back',
            request_id=eligibility_request.id
        )
        files_processed.append({
            'name': 'ID Back',
            'category': id_back_category,
            'path': id_back_path
        })
        
        # 🔥 FIX: Process Signature WITH white background correction
        print(f"\n📄 Processing Signature...")
        
        # Process signature to fix black background issue
        processed_signature = process_signature_image(signature)
        
        signature_category = smart_categorize_file(processed_signature, 'signature')
        signature_path = save_categorized_eligibility_file(
            file=processed_signature,  # Use processed signature
            category=signature_category,
            user_name=f"{first_name}_{last_name}",
            file_type='signature',
            request_id=eligibility_request.id
        )
        files_processed.append({
            'name': 'Signature',
            'category': signature_category,
            'path': signature_path
        })
        
        # Store file paths in eligibility request
        eligibility_request.id_front = id_front_path
        eligibility_request.id_back = id_back_path
        eligibility_request.signature = signature_path
        eligibility_request.save()
        
        print(f"\n{'='*80}")
        print(f"✅ SUBMISSION COMPLETE")
        print(f"📊 Files Categorized:")
        for file_info in files_processed:
            print(f"   - {file_info['name']} → {file_info['category']}")
        print(f"{'='*80}\n")
        
        # Generate reference number
        reference_number = f"EC-{timezone.now().year}-{eligibility_request.id:05d}"
        
        return JsonResponse({
            'success': True,
            'message': 'Application submitted successfully! Check your email for confirmation.',
            'id_number': reference_number,
            'request_id': eligibility_request.id,
            'files_categorized': [
                {'name': f['name'], 'category': f['category']}
                for f in files_processed
            ]
        })
        
    except Exception as e:
        print(f"\n{'='*80}")
        print(f"❌ SUBMISSION ERROR")
        print(f"{'='*80}")
        print(f"Error: {str(e)}")
        import traceback
        print(traceback.format_exc())
        print(f"{'='*80}\n")
        
        return JsonResponse({
            'success': False,
            'error': f'Submission failed: {str(e)}'
        }, status=400)



def smart_categorize_file(file, file_type_hint):
    """
    🧠 SMART CATEGORIZATION ENGINE
    Analyzes file content to determine correct folder
    
    Returns: 'appointive_certificates', 'elective_certificates', 'ids', or 'signatures'
    """
    try:
        print(f"🔍 Analyzing: {file.name}")
        
        # Force categorization based on file type hint
        if file_type_hint in ['id_front', 'id_back']:
            print(f"   ✅ Category: ids (based on file type)")
            return 'ids'
        
        if file_type_hint == 'signature':
            print(f"   ✅ Category: signatures (based on file type)")
            return 'signatures'
        
        # For other files, analyze content
        file_extension = file.name.lower().split('.')[-1]
        text_content = ""
        
        # Extract text based on file type
        if file_extension == 'pdf':
            text_content = extract_text_from_pdf(file)
        elif file_extension in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
            text_content = extract_text_from_image(file)
        
        # If we have text, analyze it
        if text_content:
            category = analyze_text_for_category(text_content, file.name)
            print(f"   ✅ Category: {category} (based on content analysis)")
            return category
        
        # Fallback to filename analysis
        category = categorize_by_filename(file.name)
        print(f"   ✅ Category: {category} (based on filename)")
        return category
        
    except Exception as e:
        print(f"   ⚠️ Categorization error: {e}")
        return 'ids'  # Default fallback


def extract_text_from_pdf(file):
    """Extract text from PDF file"""
    try:
        file.seek(0)
        pdf_reader = PyPDF2.PdfReader(file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        print(f"   ⚠️ PDF extraction error: {e}")
        return ""


def extract_text_from_image(file):
    """Extract text from image using OCR"""
    try:
        file.seek(0)
        image = Image.open(file)
        text = pytesseract.image_to_string(image)
        return text
    except Exception as e:
        print(f"   ⚠️ OCR extraction error: {e}")
        return ""


def analyze_text_for_category(text, filename):
    """
    Analyze extracted text to determine certificate type
    """
    text_lower = text.lower()
    filename_lower = filename.lower()
    
    # Keywords for different certificate types
    appointive_keywords = [
        'appointive official',
        'appointive position',
        'date of appointment',
        'appointing authority',
        'appointing punong barangay',
        'barangay secretary',
        'barangay treasurer',
        'appointment',
        'csc-erpo boe form 1(b)'
    ]
    
    elective_keywords = [
        'elective official',
        'elective position',
        'date of election',
        'term of office',
        'punong barangay',
        'sanguniang barangay member',
        'elected',
        'election',
        'csc-erpo boe form 1(a)'
    ]
    
    id_keywords = [
        'identification',
        'id card',
        'government issued id',
        'driver',
        'passport',
        'sss',
        'philhealth',
        'tin',
        'voter'
    ]
    
    signature_keywords = [
        'signature',
        'sign here',
        'e-signature'
    ]
    
    # Count keyword matches
    appointive_score = sum(1 for keyword in appointive_keywords if keyword in text_lower)
    elective_score = sum(1 for keyword in elective_keywords if keyword in text_lower)
    id_score = sum(1 for keyword in id_keywords if keyword in text_lower)
    signature_score = sum(1 for keyword in signature_keywords if keyword in text_lower)
    
    # Check filename hints
    if any(word in filename_lower for word in ['id_front', 'id_back', 'identification']):
        id_score += 5
    if 'signature' in filename_lower or 'sign' in filename_lower:
        signature_score += 5
    if 'appointive' in filename_lower:
        appointive_score += 3
    if 'elective' in filename_lower:
        elective_score += 3
    
    print(f"    Scores - Appointive:{appointive_score}, Elective:{elective_score}, ID:{id_score}, Signature:{signature_score}")
    
    # Determine category based on highest score
    scores = {
        'appointive_certificates': appointive_score,
        'elective_certificates': elective_score,
        'ids': id_score,
        'signatures': signature_score
    }
    
    max_score = max(scores.values())
    
    if max_score == 0:
        return categorize_by_filename(filename)
    
    # Return category with highest score
    for category, score in scores.items():
        if score == max_score:
            return category
    
    return 'ids'


def categorize_by_filename(filename):
    """Fallback categorization based on filename"""
    filename_lower = filename.lower()
    
    if any(word in filename_lower for word in ['id_front', 'id_back', 'identification', '_id_']):
        return 'ids'
    elif any(word in filename_lower for word in ['signature', 'sign', 'esign']):
        return 'signatures'
    elif 'appointive' in filename_lower:
        return 'appointive_certificates'
    elif 'elective' in filename_lower:
        return 'elective_certificates'
    else:
        return 'ids'


def save_categorized_eligibility_file(file, category, user_name, file_type, request_id):
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile

    file_extension = os.path.splitext(file.name)[1]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{user_name}_{file_type}_{timestamp}{file_extension}"

    year = datetime.now().strftime('%Y')
    month = datetime.now().strftime('%m')
    folder_path = f"certification_files/{category}/{year}/{month}"

    gitkeep_path = f"{folder_path}/.gitkeep"
    if not default_storage.exists(gitkeep_path):
        default_storage.save(gitkeep_path, ContentFile(b''))
    file_path = os.path.join(folder_path, filename)
    
    file.seek(0)  
    path = default_storage.save(file_path, ContentFile(file.read()))
    print(f"    ✓ Saved to: {path}")
    return path


# Add this helper function for generating certificates (when admin approves)
def generate_certificate_pdf(eligibility_request):
    """
    Generate certificate PDF with DILG logos and proper formatting
    Fixed: Logo placement, Director name, Table formatting
    """
    try:
        from django.core.files.storage import default_storage
        from django.core.files.base import ContentFile
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle
        from io import BytesIO
        import os
        from django.conf import settings
        
        print(f"\n{'='*70}")
        print(f"📄 GENERATING OFFICIAL DILG CERTIFICATE")
        print(f"{'='*70}")
        print(f"Request ID: {eligibility_request.id}")
        print(f"Name: {eligibility_request.full_name}")
        print(f"Position: {eligibility_request.position_type}")
        print(f"Barangay: {eligibility_request.barangay}")
        
        # Determine folder
        if eligibility_request.position_type == 'appointive':
            folder = 'appointive_certificates'
            form_ref = "CSC-ERPO BOE Form 1(b). April 2012"
            position_label = "(Appointive Official)"
        else:
            folder = 'elective_certificates'
            form_ref = "CSC-ERPO BOE Form 1(a) (Revised, June 2017)"
            position_label = "(Elective Official)"
        
        # Get or create category
        from .models import FileCategory, CategorizedFile
        category, _ = FileCategory.objects.get_or_create(
            name=folder,
            defaults={
                'display_name': folder.replace('_', ' ').title(),
                'folder_path': f'certification_files/{folder}/',
            }
        )
        
        # Create PDF
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # === OUTER BORDER ===
        c.setStrokeColor(colors.HexColor('#1A237E'))
        c.setLineWidth(2)
        c.rect(0.4*inch, 0.4*inch, width - 0.8*inch, height - 0.8*inch)
        
        # === INNER BORDER ===
        c.setLineWidth(0.5)
        c.rect(0.5*inch, 0.5*inch, width - 1*inch, height - 1*inch)
        
        # === DILG LOGOS (IMPROVED) ===
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'Pictures', 'logo1.png')
        
        if os.path.exists(logo_path):
            try:
                logo_size = 0.7*inch  # Increased size
                logo_y = height - 1.3*inch
                
                # Left logo
                c.drawImage(logo_path, 0.75*inch, logo_y, 
                           width=logo_size, height=logo_size, 
                           preserveAspectRatio=True, mask='auto')
                
                # Right logo
                c.drawImage(logo_path, width - 0.75*inch - logo_size, logo_y, 
                           width=logo_size, height=logo_size, 
                           preserveAspectRatio=True, mask='auto')
                print(f"✓ DILG logos added")
            except Exception as logo_err:
                print(f"⚠️ Logo error: {logo_err}")
        else:
            print(f"⚠️ Logo not found at: {logo_path}")
        
        # === HEADER ===
        y_pos = height - 1*inch
        
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 9)
        c.drawCentredString(width/2, y_pos, "Republic of the Philippines")
        
        y_pos -= 0.2*inch
        c.setFillColor(colors.HexColor('#1A237E'))
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(width/2, y_pos, "DEPARTMENT OF THE INTERIOR AND")
        y_pos -= 0.18*inch
        c.drawCentredString(width/2, y_pos, "LOCAL GOVERNMENT")
        
        y_pos -= 0.2*inch
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 9)
        c.drawCentredString(width/2, y_pos, "REGION IV-A CALABARZON")
        
        y_pos -= 0.15*inch
        c.drawCentredString(width/2, y_pos, "CITY OF LUCENA")
        
        # Form reference (right aligned)
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.gray)
        c.drawRightString(width - 0.6*inch, y_pos - 0.3*inch, form_ref)
        
        # === HORIZONTAL LINE ===
        y_pos -= 0.5*inch
        c.setStrokeColor(colors.black)
        c.setLineWidth(1)
        c.line(0.75*inch, y_pos, width - 0.75*inch, y_pos)
        
        # === TITLE SECTION ===
        y_pos -= 0.5*inch
        c.setFillColor(colors.HexColor('#1A237E'))
        c.setFont("Helvetica-Bold", 18)
        c.drawCentredString(width/2, y_pos, "CERTIFICATION")
        
        y_pos -= 0.25*inch
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 10)
        c.drawCentredString(width/2, y_pos, "on Services Rendered in the Barangay*")
        
        y_pos -= 0.2*inch
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(width/2, y_pos, position_label)
        
        # === HORIZONTAL LINE ===
        y_pos -= 0.3*inch
        c.line(0.75*inch, y_pos, width - 0.75*inch, y_pos)
        
        # === BODY TEXT ===
        y_pos -= 0.4*inch
        c.setFont("Helvetica", 10)
        c.setFillColor(colors.black)
        
        text_line = f"This is to certify that "
        c.drawString(0.75*inch, y_pos, text_line)
        
        name_x = 0.75*inch + c.stringWidth(text_line, "Helvetica", 10)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(name_x, y_pos, eligibility_request.full_name.upper())
        
        after_name_x = name_x + c.stringWidth(eligibility_request.full_name.upper(), "Helvetica-Bold", 10)
        c.setFont("Helvetica", 10)
        c.drawString(after_name_x, y_pos, " has rendered services in")
        
        y_pos -= 0.18*inch
        barangay_text = f"Barangay {eligibility_request.barangay}, with the following details:"
        c.drawString(0.75*inch, y_pos, barangay_text)
        
        # === TABLE (FIXED FORMATTING) ===
        y_pos -= 0.5*inch
        
        if eligibility_request.position_type == 'appointive':
            # Appointive table with proper spacing
            table_data = [
                ['Position\nHeld', 
                 'Date of\nAppointment', 
                 'Inclusive Dates\nFrom', 
                 'Inclusive Dates\nTo',
                 'No. of Years\nServed', 
                 'Appointing Punong\nBarangay Name',
                 'Date Elected',
                 'Term of Office\n(years)'],
                [
                    'Barangay\nSecretary',
                    eligibility_request.appointment_from.strftime('%m/%d/%Y') if eligibility_request.appointment_from else 'N/A',
                    eligibility_request.appointment_from.strftime('%m/%d/%Y') if eligibility_request.appointment_from else 'N/A',
                    eligibility_request.appointment_to.strftime('%m/%d/%Y') if eligibility_request.appointment_to else 'N/A',
                    f"{float(eligibility_request.years_in_service)} yrs" if eligibility_request.years_in_service else '0.0 yrs',
                    eligibility_request.appointing_punong_barangay or 'N/A',
                    eligibility_request.pb_date_elected.strftime('%m/%d/%Y') if eligibility_request.pb_date_elected else 'N/A',
                    f"{float(eligibility_request.pb_years_service)} yrs" if eligibility_request.pb_years_service else '0.0 yrs'
                ]
            ]
            
            # Adjusted column widths for better spacing
            col_widths = [0.8*inch, 0.75*inch, 0.75*inch, 0.75*inch, 0.7*inch, 1.1*inch, 0.7*inch, 0.7*inch]
            row_heights = [0.6*inch, 0.5*inch]
            
        else:
            # Elective table - matching official format exactly
            table_data = [
                ['Position Held', 'Date of Election\n(mm/dd/yyyy)', 'Term of Office\n(no. of years)', 
                 'Inclusive Dates\nFrom (mm/dd/yyyy)', 'Inclusive Dates\nTo (mm/dd/yyyy)'],
                [
                    eligibility_request.position_held or 'Punong Barangay',
                    eligibility_request.election_from.strftime('%m/%d/%Y') if eligibility_request.election_from else '',
                    eligibility_request.term_office or 'November 2025 -\nNovember 2025',
                    eligibility_request.election_from.strftime('%m/%d/%Y') if eligibility_request.election_from else '',
                    eligibility_request.election_to.strftime('%m/%d/%Y') if eligibility_request.election_to else ''
                ]
            ]
            
            col_widths = [1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch]
            row_heights = [0.6*inch, 0.5*inch]
        
        # Create table with improved styling
        table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
        
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A237E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            
            # Data styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            
            # Borders
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            
            # Padding
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        table.wrapOn(c, width, height)
        table_height = table._height
        table.drawOn(c, 0.75*inch, y_pos - table_height)
        
        y_pos -= (table_height + 0.4*inch)
        
        # === COMPLETED TERM SECTION (ELECTIVE ONLY) ===
        if eligibility_request.position_type == 'elective':
            c.setFont("Helvetica-Bold", 9)
            c.setFillColor(colors.black)
            c.drawString(0.75*inch, y_pos, "Completed Term of Office?")
            c.setFont("Helvetica", 9)
            c.drawString(2.5*inch, y_pos, "(Please check (√) appropriate box)")
            
            y_pos -= 0.25*inch
            
            completed_term = eligibility_request.completed_term
            checkbox_size = 0.12*inch
            
            # YES checkbox
            c.rect(0.95*inch, y_pos - checkbox_size/2, checkbox_size, checkbox_size)
            if completed_term and completed_term.lower() == 'yes':
                # Draw checkmark
                c.setLineWidth(2)
                c.line(0.97*inch, y_pos - checkbox_size/4, 
                       1.0*inch, y_pos - checkbox_size/1.5)
                c.line(1.0*inch, y_pos - checkbox_size/1.5,
                       1.05*inch, y_pos)
                c.setLineWidth(0.5)
            
            c.drawString(1.15*inch, y_pos - 0.05*inch, "YES")
            
            # NO checkbox
            c.rect(1.8*inch, y_pos - checkbox_size/2, checkbox_size, checkbox_size)
            if completed_term and completed_term.lower() == 'no':
                # Draw checkmark
                c.setLineWidth(2)
                c.line(1.82*inch, y_pos - checkbox_size/4, 
                       1.85*inch, y_pos - checkbox_size/1.5)
                c.line(1.85*inch, y_pos - checkbox_size/1.5,
                       1.90*inch, y_pos)
                c.setLineWidth(0.5)
            
            c.drawString(2.0*inch, y_pos - 0.05*inch, "NO. Specify total number of days not served")
            
            y_pos -= 0.25*inch
            
            # If NO was selected, show reason box
            if completed_term and completed_term.lower() == 'no':
                c.setFont("Helvetica-Bold", 9)
                c.drawString(0.95*inch, y_pos, "Reason for non-completion:")
                
                y_pos -= 0.25*inch
                
                # Draw reason box
                reason_text = eligibility_request.incomplete_reason or 'Not specified'
                c.setFont("Helvetica", 9)
                c.setFillColor(colors.HexColor('#f5f5f5'))
                c.rect(0.95*inch, y_pos - 0.35*inch, 5.5*inch, 0.4*inch, fill=1)
                
                c.setFillColor(colors.black)
                c.drawString(1.05*inch, y_pos - 0.15*inch, reason_text)
                
                y_pos -= 0.45*inch
            
            # ASSUMED checkbox (always show)
            c.rect(0.95*inch, y_pos - checkbox_size/2, checkbox_size, checkbox_size)
            c.drawString(1.15*inch, y_pos - 0.05*inch, "Assumed under rule on succession.")
            
            y_pos -= 0.35*inch
        
        # === FOOTER TEXT ===
        c.setFont("Helvetica", 9)
        c.setFillColor(colors.black)
        
        if eligibility_request.position_type == 'elective':
            # Elective footer text with CSC Resolutions
            footer_text = f"This Certification is issued in support of the evaluation/processing of the application of {eligibility_request.full_name.upper()}"
            c.drawString(0.75*inch, y_pos, footer_text)
            
            y_pos -= 0.15*inch
            footer_text2 = "for the grant of Barangay Official Eligibility pursuant to Republic Act No. 7160, in accordance"
            c.drawString(0.75*inch, y_pos, footer_text2)
            
            y_pos -= 0.15*inch
            footer_text3 = "with CSC Resolution No. 1200865 dated June 14, 2012 and CSC Resolution No."
            c.drawString(0.75*inch, y_pos, footer_text3)
            
            y_pos -= 0.15*inch
            footer_text4 = "1601257 dated November 21, 2016."
            c.drawString(0.75*inch, y_pos, footer_text4)
        else:
            # Appointive footer text
            footer_text = f"This Certification is issued in support of the evaluation/processing of the application of {eligibility_request.full_name.upper()}"
            c.drawString(0.75*inch, y_pos, footer_text)
            
            y_pos -= 0.15*inch
            footer_text2 = "for the grant of Barangay Official Eligibility pursuant to Republic Act No. 7160, in accordance"
            c.drawString(0.75*inch, y_pos, footer_text2)
            
            y_pos -= 0.15*inch
            footer_text3 = "with CSC Resolution No. 13 series of 2012."
            c.drawString(0.75*inch, y_pos, footer_text3)
        
        # === SIGNATURE SECTION ===
        y_pos -= 0.5*inch
        
        from django.utils import timezone
        date_text = f"Lucena City, Quezon, {timezone.now().strftime('%B %d, %Y')}."
        c.drawString(0.75*inch, y_pos, date_text)
        
        # === SIGNATURE SECTION (UPDATED DIRECTOR) ===
        y_pos -= 0.8*inch
        
        # Signature line (right side)
        sig_x = width - 2.8*inch
        c.setStrokeColor(colors.black)
        c.setLineWidth(0.5)
        c.line(sig_x, y_pos, width - 0.75*inch, y_pos)
        
        # Director name and title
        y_pos -= 0.2*inch
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.black)
        c.drawCentredString(sig_x + 1.025*inch, y_pos, "LEANDRO SIPOY GIGANTOCA, CESE")
        
        y_pos -= 0.18*inch
        c.setFont("Helvetica", 9)
        c.drawCentredString(sig_x + 1.025*inch, y_pos, "OIC-HUC Director, Lucena City")
        
        # Save PDF
        c.showPage()
        c.save()
        
        pdf_data = buffer.getvalue()
        buffer.close()
        
        print(f"✓ PDF generated successfully ({len(pdf_data)} bytes)")
        
        # Create filename
        safe_name = eligibility_request.full_name.replace(' ', '_').replace('.', '')
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{safe_name}_{eligibility_request.position_type}_Certificate_{timestamp}.pdf"
        
        # Save to storage
        file_path = f"certification_files/{folder}/{filename}"
        saved_path = default_storage.save(file_path, ContentFile(pdf_data))
        
        print(f"✓ PDF saved: {saved_path}")
        
        # Create CategorizedFile entry
        categorized = CategorizedFile.objects.create(
            file=saved_path,
            original_filename=filename,
            file_type='pdf',
            file_size=len(pdf_data),
            mime_type='application/pdf',
            category=category,
            source='eligibility',
            detected_content=f'{eligibility_request.get_position_type_display()} Certificate',
            eligibility_request=eligibility_request,
            uploaded_by=eligibility_request.approved_by,
            tags=f"{eligibility_request.full_name}, {eligibility_request.position_type}, Certificate, Approved"
        )
        
        print(f"✓ CategorizedFile created: ID {categorized.id}")
        category.update_file_count()
        print(f"✓ Updated folder file count: {category.file_count}")
        
        print(f"{'='*70}\n")
        
        return saved_path
        
    except Exception as e:
        print(f"\n{'='*70}")
        print(f"❌ CERTIFICATE GENERATION ERROR")
        print(f"{'='*70}")
        print(f"Error: {str(e)}")
        import traceback
        print(traceback.format_exc())
        print(f"{'='*70}\n")
        return None


# Add these API endpoints for the certificate files page

@require_http_methods(["GET"])
def get_certificate_files_by_category(request, category):
    """
    FIXED: Get certificate files by category with REAL database IDs
    """
    try:
        from django.core.files.storage import default_storage
        from .models import CategorizedFile, FileCategory
        
        print(f"\n{'='*70}")
        print(f"📁 GET CERTIFICATE FILES BY CATEGORY")
        print(f"{'='*70}")
        print(f"Category requested: {category}")
        
        # Map category names to folder paths
        category_folders = {
            'certificates': ['appointive_certificates', 'elective_certificates'],
            'appointive_certificates': ['appointive_certificates'],
            'elective_certificates': ['elective_certificates'],
            'ids': ['ids'],
            'signatures': ['signatures']
        }
        
        if category not in category_folders:
            return JsonResponse({
                'success': False,
                'error': f'Invalid category: {category}'
            }, status=400)
        
        folders_to_scan = category_folders[category]
        all_files = []
        
        # Get files from CategorizedFile database
        for folder_name in folders_to_scan:
            print(f"\n🔍 Querying CategorizedFile for: {folder_name}")
            
            # Get or create category
            try:
                file_category = FileCategory.objects.get(name=folder_name)
            except FileCategory.DoesNotExist:
                print(f"   ⚠️ Category '{folder_name}' not found in database")
                continue
            
            # Query CategorizedFile with REAL IDs
            files = CategorizedFile.objects.filter(
                category=file_category
            ).select_related('uploaded_by', 'barangay')
            
            print(f"   ✓ Found {files.count()} files in database")
            
            for file_obj in files:
                try:
                    file_info = {
                        'id': file_obj.id,  # ✅ REAL DATABASE ID
                        'filename': file_obj.original_filename,
                        'file_url': file_obj.file.url if file_obj.file else '',
                        'file_type': file_obj.file_type,
                        'file_size': file_obj.file_size_mb,
                        'uploaded_at': file_obj.uploaded_at.strftime('%B %d, %Y %I:%M %p'),
                        'category': file_category.display_name,
                        'folder': folder_name,
                        'uploaded_by': (
                            file_obj.uploaded_by.get_full_name() 
                            if file_obj.uploaded_by 
                            else 'System'
                        ),
                        'barangay': (
                            file_obj.barangay.name 
                            if file_obj.barangay 
                            else None
                        )
                    }
                    
                    all_files.append(file_info)
                    print(f"    ✓ {file_obj.original_filename} (ID: {file_obj.id})")
                    
                except Exception as file_err:
                    print(f"    ✗ Error processing file {file_obj.id}: {file_err}")
                    continue
        
        print(f"\n📊 Total files found: {len(all_files)}")
        print(f"{'='*70}\n")
        
        return JsonResponse({
            'success': True,
            'files': all_files,
            'total_count': len(all_files),
            'category': category
        })
        
    except Exception as e:
        print(f"\n{'='*70}")
        print(f"❌ ERROR: {str(e)}")
        print(f"{'='*70}\n")
        import traceback
        print(traceback.format_exc())
        
        return JsonResponse({
            'success': False,
            'error': str(e),
            'category': category
        }, status=500)


def save_categorized_eligibility_file(file, category, user_name, file_type, request_id):
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile
    from .models import FileCategory, CategorizedFile

    file_extension = os.path.splitext(file.name)[1]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{user_name}_{file_type}_{timestamp}{file_extension}"

    year = datetime.now().strftime('%Y')
    month = datetime.now().strftime('%m')
    folder_path = f"certification_files/{category}/{year}/{month}"

    # Ensure .gitkeep exists
    gitkeep_path = f"{folder_path}/.gitkeep"
    if not default_storage.exists(gitkeep_path):
        default_storage.save(gitkeep_path, ContentFile(b''))
    
    # Save file
    file_path = os.path.join(folder_path, filename)
    file.seek(0)  
    path = default_storage.save(file_path, ContentFile(file.read()))
    print(f"    ✓ Saved to: {path}")
    
    # ✅ FIX: Create/get FileCategory and CategorizedFile
    file_category, _ = FileCategory.objects.get_or_create(
        name=category,  # 'ids' or 'signatures'
        defaults={
            'display_name': category.replace('_', ' ').title(),
            'folder_path': f'certification_files/{category}/'
        }
    )
    
    # Create CategorizedFile entry
    categorized = CategorizedFile.objects.create(
        file=path,
        original_filename=filename,
        file_type='image' if file_extension.lower() in ['.jpg', '.jpeg', '.png', '.gif'] else 'document',
        file_size=file.size,
        mime_type=file.content_type,
        category=file_category,
        source='eligibility',
        detected_content=f'ID {file_type}' if 'id' in file_type else 'Signature',
        tags=f"{user_name}, {file_type}, Eligibility Request {request_id}"
    )
    
    # Update category file count
    file_category.update_file_count()
    
    print(f"    ✓ Created CategorizedFile ID: {categorized.id}")
    
    return path

@require_http_methods(["GET"])
def debug_certificate_categories(request):
    """Debug view to see what's in the database"""
    from .models import FileCategory, CategorizedFile
    
    result = {
        'file_categories': [],
        'files_by_folder': {},
        'total_files': CategorizedFile.objects.count()
    }
    
    # Get all FileCategory objects
    for cat in FileCategory.objects.all():
        result['file_categories'].append({
            'id': cat.id,
            'name': cat.name,
            'display_name': cat.display_name,
            'file_count': CategorizedFile.objects.filter(category=cat).count()
        })
    
    # Get all CategorizedFile objects grouped by their source
    files = CategorizedFile.objects.select_related('category').all()
    
    for file_obj in files:
        folder = file_obj.category.name if file_obj.category else 'No Category'
        
        if folder not in result['files_by_folder']:
            result['files_by_folder'][folder] = []
        
        result['files_by_folder'][folder].append({
            'id': file_obj.id,
            'filename': file_obj.original_filename,
            'source': file_obj.source,
            'file_type': file_obj.file_type,
        })
    
    return JsonResponse(result, json_dumps_params={'indent': 2})


@require_http_methods(["GET"])
def debug_certificate_files(request):
    from django.core.files.storage import default_storage
    import os
    
    try:
        result = {
            'base_path': 'certification_files',
            'folders': {},
            'all_files_found': []
        }
        
        # Check each category folder
        for category in ['appointive_certificates', 'elective_certificates', 'ids', 'signatures']:
            folder_path = f"certification_files/{category}/"
            
            folder_info = {
                'exists': default_storage.exists(folder_path),
                'files_found': []
            }
            
            if folder_info['exists']:
                try:
                    # Recursively find all files
                    def scan_directory(path):
                        files_in_dir = []
                        try:
                            dirs, files = default_storage.listdir(path)
                            
                            # Add files in current directory
                            for f in files:
                                if not f.startswith('.'):
                                    full_path = os.path.join(path, f)
                                    files_in_dir.append({
                                        'path': full_path,
                                        'name': f,
                                        'url': default_storage.url(full_path),
                                        'size': default_storage.size(full_path)
                                    })
                            
                            # Recursively scan subdirectories
                            for d in dirs:
                                subdir_path = os.path.join(path, d)
                                files_in_dir.extend(scan_directory(subdir_path))
                        
                        except Exception as e:
                            print(f"Error scanning {path}: {e}")
                        
                        return files_in_dir
                    
                    folder_info['files_found'] = scan_directory(folder_path)
                    result['all_files_found'].extend(folder_info['files_found'])
                    
                except Exception as e:
                    folder_info['error'] = str(e)
            
            result['folders'][category] = folder_info
        
        # Summary
        result['summary'] = {
            'total_files': len(result['all_files_found']),
            'files_by_category': {
                cat: len(info['files_found']) 
                for cat, info in result['folders'].items()
            }
        }
        
        return JsonResponse(result, json_dumps_params={'indent': 2})
        
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)

def test_certificate_setup(request):

    try:
        from django.core.files.storage import default_storage
        import os
        
        results = {
            'status': 'Testing Certificate Files Setup',
            'media_root': settings.MEDIA_ROOT if hasattr(settings, 'MEDIA_ROOT') else 'Not configured',
            'media_url': settings.MEDIA_URL if hasattr(settings, 'MEDIA_URL') else 'Not configured',
            'folders': {},
            'sample_structure': {}
        }
        
        # Check each category folder
        categories = ['appointive_certificates', 'elective_certificates', 'ids', 'signatures']
        
        for category in categories:
            folder_path = f'certification_files/{category}/'
            
            folder_info = {
                'exists': default_storage.exists(folder_path),
                'full_path': os.path.join(settings.MEDIA_ROOT, folder_path) if hasattr(settings, 'MEDIA_ROOT') else 'Unknown',
                'file_count': 0,
                'sample_files': []
            }
            
            if folder_info['exists']:
                try:
                    # Try to count files
                    year_dirs, direct_files = default_storage.listdir(folder_path)
                    folder_info['file_count'] = len(direct_files)
                    folder_info['sample_files'] = direct_files[:3]  # First 3 files
                    folder_info['year_folders'] = year_dirs
                except Exception as e:
                    folder_info['error'] = str(e)
            
            results['folders'][category] = folder_info
        
        # Show expected structure
        results['sample_structure'] = {
            'certification_files/': {
                'appointive_certificates/': {
                    '2024/': {
                        '11/': ['John_Doe_certificate_20241106.pdf']
                    }
                },
                'elective_certificates/': {
                    '2024/': {
                        '11/': ['Jane_Smith_certificate_20241106.pdf']
                    }
                },
                'ids/': {
                    '2024/': {
                        '11/': ['John_Doe_id_front_20241106.jpg', 'John_Doe_id_back_20241106.jpg']
                    }
                },
                'signatures/': {
                    '2024/': {
                        '11/': ['John_Doe_signature_20241106.png']
                    }
                }
            }
        }
        
        # Test URL patterns
        from django.urls import reverse, NoReverseMatch
        
        url_tests = {}
        for category in categories:
            try:
                url = reverse('get_certificate_files_by_category', kwargs={'category': category})
                url_tests[category] = {
                    'url': url,
                    'status': 'URL pattern exists ✓'
                }
            except NoReverseMatch:
                url_tests[category] = {
                    'status': 'URL pattern MISSING ✗',
                    'fix': f"Add path('api/certificate-files/category/<str:category>/', views.get_certificate_files_by_category) to urls.py"
                }
        
        results['url_patterns'] = url_tests
        
        return JsonResponse(results, json_dumps_params={'indent': 2})
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500, json_dumps_params={'indent': 2})


@login_required
def application_request(request):
    # Get all eligibility requests
    requests = EligibilityRequest.objects.all().order_by('-date_submitted')
    
    context = {
        'requests': requests
    }
    return render(request, 'application_request.html', context)



@login_required
@require_http_methods(["POST"])
def update_application_status(request):
    """Update status AND generate certificate when approved"""
    try:
        data = json.loads(request.body)
        request_id = data.get('id')
        new_status = data.get('status')
        
        eligibility_request = get_object_or_404(EligibilityRequest, id=request_id)
        
        # Update status
        old_status = eligibility_request.status
        eligibility_request.status = new_status
        
        if new_status in ['approved', 'rejected']:
            eligibility_request.approved_by = request.user
            eligibility_request.date_processed = timezone.now()
        
        eligibility_request.save()
        
        #  GENERATE CERTIFICATE when status changes to approved
        certificate_path = None
        if new_status == 'approved' and old_status != 'approved':
            print(f"\n APPROVAL DETECTED - Generating certificate...")
            certificate_path = generate_certificate_pdf(eligibility_request)
            
            if certificate_path:
                print(f" Certificate generated: {certificate_path}")
            else:
                print(f" Certificate generation failed!")
        
        response_data = {
            'success': True,
            'message': f'Status updated to {new_status.capitalize()}',
            'new_status': new_status
        }
        
        if certificate_path:
            import os
            response_data['certificate_generated'] = True
            response_data['certificate_filename'] = os.path.basename(certificate_path)
        
        return JsonResponse(response_data)
        
    except Exception as e:
        import traceback
        print(f"❌ Error: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



def generate_certificate_pdf(eligibility_request):
    """
    Generate certificate PDF - FIXED for elective incomplete term
    """
    try:
        from django.core.files.storage import default_storage
        from django.core.files.base import ContentFile
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle
        from io import BytesIO
        import os
        from django.conf import settings
        
        print(f"\n{'='*70}")
        print(f"📄 GENERATING CERTIFICATE")
        print(f"Position: {eligibility_request.position_type}")
        if eligibility_request.position_type == 'elective':
            print(f"Completed Term: {eligibility_request.completed_term}")
            print(f"Days Not Served: {eligibility_request.days_not_served}")
            print(f"Reason: {eligibility_request.incomplete_reason}")
        print(f"{'='*70}")
        
        # Determine folder based on position type AND completion status
        if eligibility_request.position_type == 'appointive':
            folder = 'appointive_certificates'
            form_ref = "CSC-ERPO BOE Form 1(b). April 2012"
            position_label = "(Appointive Official)"
        else:
            folder = 'elective_certificates'
            form_ref = "CSC-ERPO BOE Form 1(a) (Revised, June 2017)"
            position_label = "(Elective Official)"
        
        from .models import FileCategory, CategorizedFile
        category, _ = FileCategory.objects.get_or_create(
            name=folder,
            defaults={
                'display_name': folder.replace('_', ' ').title(),
                'folder_path': f'certification_files/{folder}/',
            }
        )
        
        # Create PDF
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # === BORDERS ===
        c.setStrokeColor(colors.HexColor('#1A237E'))
        c.setLineWidth(2)
        c.rect(0.4*inch, 0.4*inch, width - 0.8*inch, height - 0.8*inch)
        
        c.setLineWidth(0.5)
        c.rect(0.5*inch, 0.5*inch, width - 1*inch, height - 1*inch)
        
        # === LOGOS ===
        from reportlab.lib.utils import ImageReader
        
        # Try multiple possible logo paths
        possible_logo_paths = [
            os.path.join(settings.BASE_DIR, 'static', 'Pictures', 'logo1.png'),
            os.path.join(settings.BASE_DIR, 'app', 'static', 'Pictures', 'logo1.png'),
            os.path.join(settings.BASE_DIR, 'static', 'pictures', 'logo1.png'),
            os.path.join(settings.STATIC_ROOT, 'Pictures', 'logo1.png') if hasattr(settings, 'STATIC_ROOT') and settings.STATIC_ROOT else None,
        ]
        
        logo_path = None
        for path in possible_logo_paths:
            if path and os.path.exists(path):
                logo_path = path
                print(f"✓ Logo found at: {path}")
                break
        
        if logo_path:
            try:
                # Use ImageReader for better compatibility
                img = ImageReader(logo_path)
                logo_size = 0.7*inch
                logo_y = height - 1.3*inch
                
                # Draw left logo
                c.drawImage(img, 0.75*inch, logo_y, 
                           width=logo_size, height=logo_size, 
                           preserveAspectRatio=True, mask='auto')
                
                # Draw right logo (reload for second instance)
                img2 = ImageReader(logo_path)
                c.drawImage(img2, width - 0.75*inch - logo_size, logo_y, 
                           width=logo_size, height=logo_size, 
                           preserveAspectRatio=True, mask='auto')
                
                print(f"✓ Logos rendered successfully on both sides")
            except Exception as e:
                print(f"⚠️ Logo rendering error: {e}")
                import traceback
                print(traceback.format_exc())
        else:
            print(f"⚠️ Logo file not found. Searched:")
            for path in possible_logo_paths:
                if path:
                    print(f"   - {path}")
        
        # === HEADER ===
        y_pos = height - 1*inch
        
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 9)
        c.drawCentredString(width/2, y_pos, "Republic of the Philippines")
        
        y_pos -= 0.2*inch
        c.setFillColor(colors.HexColor('#1A237E'))
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(width/2, y_pos, "DEPARTMENT OF THE INTERIOR AND")
        y_pos -= 0.18*inch
        c.drawCentredString(width/2, y_pos, "LOCAL GOVERNMENT")
        
        y_pos -= 0.2*inch
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 9)
        c.drawCentredString(width/2, y_pos, "REGION IV-A CALABARZON")
        
        y_pos -= 0.15*inch
        c.drawCentredString(width/2, y_pos, "CITY OF LUCENA")
        
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.gray)
        c.drawRightString(width - 0.6*inch, y_pos - 0.3*inch, form_ref)
        
        # === LINE ===
        y_pos -= 0.5*inch
        c.setStrokeColor(colors.black)
        c.setLineWidth(1)
        c.line(0.75*inch, y_pos, width - 0.75*inch, y_pos)
        
        # === TITLE ===
        y_pos -= 0.5*inch
        c.setFillColor(colors.HexColor('#1A237E'))
        c.setFont("Helvetica-Bold", 18)
        c.drawCentredString(width/2, y_pos, "CERTIFICATION")
        
        y_pos -= 0.25*inch
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 10)
        c.drawCentredString(width/2, y_pos, "on Services Rendered in the Barangay*")
        
        y_pos -= 0.2*inch
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(width/2, y_pos, position_label)
        
        # === LINE ===
        y_pos -= 0.3*inch
        c.line(0.75*inch, y_pos, width - 0.75*inch, y_pos)
        
        # === BODY ===
        y_pos -= 0.4*inch
        c.setFont("Helvetica", 10)
        c.setFillColor(colors.black)
        
        # First line with indentation (0.5 inch indent)
        indent = 0.5*inch
        text_line = f"This is to certify that "
        c.drawString(0.75*inch + indent, y_pos, text_line)
        
        name_x = 0.75*inch + indent + c.stringWidth(text_line, "Helvetica", 10)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(name_x, y_pos, eligibility_request.full_name.upper())
        
        after_name_x = name_x + c.stringWidth(eligibility_request.full_name.upper(), "Helvetica-Bold", 10)
        c.setFont("Helvetica", 10)
        c.drawString(after_name_x, y_pos, " has rendered services in")
        
        # Second line - no indent
        y_pos -= 0.18*inch
        barangay_text = f"Barangay {eligibility_request.barangay}, with the following details:"
        c.drawString(0.75*inch, y_pos, barangay_text)
        
        # === TABLE ===
        y_pos -= 0.5*inch
        
        if eligibility_request.position_type == 'elective':
            # ✅ ELECTIVE TABLE
            table_data = [
                ['Position Held', 'Date of Election\n(mm/dd/yyyy)', 'Term of Office\n(no. of years)', 
                 'From\n(mm/dd/yyyy)', 'To\n(mm/dd/yyyy)'],
                [
                    eligibility_request.position_held or 'N/A',
                    eligibility_request.election_from.strftime('%m/%d/%Y') if eligibility_request.election_from else 'N/A',
                    eligibility_request.term_office or 'N/A',
                    eligibility_request.election_from.strftime('%m/%d/%Y') if eligibility_request.election_from else 'N/A',
                    eligibility_request.election_to.strftime('%m/%d/%Y') if eligibility_request.election_to else 'N/A'
                ]
            ]
            
            col_widths = [1.4*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch]
            row_heights = [0.5*inch, 0.4*inch]
            
        else:
            # APPOINTIVE TABLE
            table_data = [
                ['Position\nHeld', 'Date of\nAppointment', 'Inclusive Dates\nFrom', 'Inclusive Dates\nTo',
                 'No. of Years\nServed', 'Appointing Punong\nBarangay Name', 'Date Elected', 'Term of Office\n(years)'],
                [
                    'Barangay\nSecretary',
                    eligibility_request.appointment_from.strftime('%m/%d/%Y') if eligibility_request.appointment_from else 'N/A',
                    eligibility_request.appointment_from.strftime('%m/%d/%Y') if eligibility_request.appointment_from else 'N/A',
                    eligibility_request.appointment_to.strftime('%m/%d/%Y') if eligibility_request.appointment_to else 'N/A',
                    f"{float(eligibility_request.years_in_service)} yrs" if eligibility_request.years_in_service else '0.0 yrs',
                    eligibility_request.appointing_punong_barangay or 'N/A',
                    eligibility_request.pb_date_elected.strftime('%m/%d/%Y') if eligibility_request.pb_date_elected else 'N/A',
                    f"{float(eligibility_request.pb_years_service)} yrs" if eligibility_request.pb_years_service else '0.0 yrs'
                ]
            ]
            col_widths = [0.8*inch, 0.75*inch, 0.75*inch, 0.75*inch, 0.7*inch, 1.1*inch, 0.7*inch, 0.7*inch]
            row_heights = [0.6*inch, 0.5*inch]
        
        table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A237E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        table.wrapOn(c, width, height)
        table_height = table._height
        table.drawOn(c, 0.75*inch, y_pos - table_height)
        
        y_pos -= (table_height + 0.3*inch)
        
        # === ✅ COMPLETED TERM SECTION (ELECTIVE ONLY) ===
        if eligibility_request.position_type == 'elective':
            c.setFont("Helvetica-Bold", 10)
            c.setFillColor(colors.black)
            c.drawString(0.75*inch, y_pos, "Completed Term of Office?")
            
            c.setFont("Helvetica", 9)
            c.drawString(2.9*inch, y_pos, "(Please check (√) appropriate box)")
            
            y_pos -= 0.3*inch
            
            completed_term = eligibility_request.completed_term
            checkbox_size = 0.15*inch
            
            # YES checkbox
            c.setStrokeColor(colors.black)
            c.setLineWidth(1)
            c.rect(1.0*inch, y_pos - 0.05*inch, checkbox_size, checkbox_size)
            
            if completed_term and completed_term.lower() == 'yes':
                c.setFont("Helvetica-Bold", 14)
                c.drawString(1.03*inch, y_pos - 0.02*inch, "✓")
            
            c.setFont("Helvetica", 10)
            c.drawString(1.25*inch, y_pos, "YES")
            
            # NO checkbox
            c.rect(2.1*inch, y_pos - 0.05*inch, checkbox_size, checkbox_size)
            
            if completed_term and completed_term.lower() == 'no':
                c.setFont("Helvetica-Bold", 14)
                c.drawString(2.13*inch, y_pos - 0.02*inch, "✓")
            
            c.setFont("Helvetica", 10)
            c.drawString(2.35*inch, y_pos, "NO, Specify total number of days not served")
            
            y_pos -= 0.3*inch
            
            # REASON BOX - ONLY SHOWS WHEN completed_term == 'no'
            if completed_term and completed_term.lower() == 'no':
                c.setFont("Helvetica-Bold", 9)
                c.drawString(1.0*inch, y_pos, "Reason for non-completion:")
                
                y_pos -= 0.25*inch
                
                c.setFillColor(colors.HexColor('#f5f5f5'))
                c.setStrokeColor(colors.HexColor('#cccccc'))
                c.setLineWidth(0.5)
                reason_box_height = 0.7*inch
                c.rect(1.0*inch, y_pos - reason_box_height, 5.5*inch, reason_box_height, fill=1, stroke=1)
                
                c.setFillColor(colors.black)
                c.setFont("Helvetica", 9)
                
                reason_text = eligibility_request.incomplete_reason or 'Not specified'
                max_width = 5.2*inch
                
                words = reason_text.split()
                lines = []
                current_line = []
                
                for word in words:
                    test_line = ' '.join(current_line + [word])
                    if c.stringWidth(test_line, "Helvetica", 9) < max_width:
                        current_line.append(word)
                    else:
                        if current_line:
                            lines.append(' '.join(current_line))
                        current_line = [word]
                
                if current_line:
                    lines.append(' '.join(current_line))
                
                text_y = y_pos - 0.2*inch
                for i, line in enumerate(lines[:4]):
                    c.drawString(1.1*inch, text_y, line)
                    text_y -= 0.13*inch
                
                y_pos -= (reason_box_height + 0.2*inch)
            
            # "Assumed under rule on succession" checkbox
            y_pos -= 0.25*inch
            c.setStrokeColor(colors.black)
            c.rect(1.0*inch, y_pos - 0.05*inch, checkbox_size, checkbox_size)
            c.setFont("Helvetica", 9)
            c.drawString(1.25*inch, y_pos, "Assumed under rule on succession.")
            
            y_pos -= 0.35*inch
        
        # === FOOTER TEXT ===
        c.setFont("Helvetica", 9)
        c.setFillColor(colors.black)
        
        # Create properly wrapped footer text with indentation
        left_margin = 0.75*inch
        indent = 0.5*inch  # Same indent as opening paragraph
        max_width = width - 1.5*inch - indent  # Account for indent
        
        if eligibility_request.position_type == 'elective':
            footer_text = f"This Certification is issued in support of the evaluation/processing of the application of {eligibility_request.full_name.upper()} for the grant of Barangay Official Eligibility pursuant to Republic Act No. 7160, in accordance with CSC Resolution No. 1200865 dated June 14, 2012 and CSC Resolution No. 1601257 dated November 21, 2016."
        else:
            footer_text = f"This Certification is issued in support of the evaluation/processing of the application of {eligibility_request.full_name.upper()} for the grant of Barangay Official Eligibility pursuant to Republic Act No. 7160, in accordance with CSC Resolution No. 13 series of 2012."
        
        # Word wrap the footer text
        words = footer_text.split()
        lines = []
        current_line = []
        
        for word in words:
            test_line = ' '.join(current_line + [word])
            if c.stringWidth(test_line, "Helvetica", 9) < max_width:
                current_line.append(word)
            else:
                if current_line:
                    lines.append(' '.join(current_line))
                current_line = [word]
        
        if current_line:
            lines.append(' '.join(current_line))
        
        # Draw wrapped footer text with indentation on first line only
        for i, line in enumerate(lines):
            if i == 0:
                # First line gets indentation
                c.drawString(left_margin + indent, y_pos, line)
            else:
                # Subsequent lines align with left margin
                c.drawString(left_margin, y_pos, line)
            y_pos -= 0.15*inch
        
        # === DATE ===
        y_pos -= 0.2*inch
        from django.utils import timezone
        date_text = f"Lucena City, Quezon, {timezone.now().strftime('%B %d, %Y')}."
        c.drawString(0.75*inch, y_pos, date_text)
        
        # === SIGNATURES ===
        y_pos -= 0.8*inch
        
        # Director signature (RIGHT SIDE - properly aligned)
        sig_line_width = 2.5*inch
        sig_right_margin = 0.75*inch
        sig_line_x_start = width - sig_right_margin - sig_line_width
        sig_line_x_end = width - sig_right_margin
        
        # Draw signature line
        c.setStrokeColor(colors.black)
        c.setLineWidth(0.5)
        c.line(sig_line_x_start, y_pos, sig_line_x_end, y_pos)
        
        # Director name (centered above line)
        y_pos -= 0.18*inch
        sig_center_x = sig_line_x_start + (sig_line_width / 2)
        
        c.setFont("Helvetica-Bold", 10)
        name_width = c.stringWidth("LEANDRO SIPOY GIGANTOCA, CESE", "Helvetica-Bold", 10)
        c.drawString(sig_center_x - (name_width / 2), y_pos, "LEANDRO SIPOY GIGANTOCA, CESE")
        
        # Director title (centered below name)
        y_pos -= 0.15*inch
        c.setFont("Helvetica", 9)
        title_width = c.stringWidth("OIC-HUC Director, Lucena City", "Helvetica", 9)
        c.drawString(sig_center_x - (title_width / 2), y_pos, "OIC-HUC Director, Lucena City")
        
        # Save PDF
        c.showPage()
        c.save()
        
        pdf_data = buffer.getvalue()
        buffer.close()
        
        print(f"✓ PDF generated ({len(pdf_data)} bytes)")
        
        # ✅ FIXED FILENAME GENERATION
        safe_name = eligibility_request.full_name.replace(' ', '_').replace('.', '')
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        
        if eligibility_request.position_type == 'elective':
            completion_status = 'Completed' if eligibility_request.completed_term and eligibility_request.completed_term.lower() == 'yes' else 'Incomplete'
            filename = f"{safe_name}_Elective_{completion_status}_Certificate_{timestamp}.pdf"
        else:
            filename = f"{safe_name}_Appointive_Certificate_{timestamp}.pdf"
        
        # Save to storage
        file_path = f"certification_files/{folder}/{filename}"
        saved_path = default_storage.save(file_path, ContentFile(pdf_data))
        
        print(f"✓ Saved: {saved_path}")
        
        # Create CategorizedFile record
        completion_tag = 'Completed' if eligibility_request.position_type == 'elective' and eligibility_request.completed_term and eligibility_request.completed_term.lower() == 'yes' else 'Incomplete'
        
        categorized = CategorizedFile.objects.create(
            file=saved_path,
            original_filename=filename,
            file_type='pdf',
            file_size=len(pdf_data),
            mime_type='application/pdf',
            category=category,
            source='eligibility',
            detected_content=f'{eligibility_request.get_position_type_display()} Certificate - {completion_tag}',
            eligibility_request=eligibility_request,
            uploaded_by=eligibility_request.approved_by,
            tags=f"{eligibility_request.full_name}, {eligibility_request.position_type}, {completion_tag}"
        )
        
        category.update_file_count()
        print(f"{'='*70}\n")
        
        return saved_path
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return None


# Add to views.py
from PIL import Image, ImageOps
import io

def process_signature_image(uploaded_file):
    """
    Process signature image to ensure it has white background
    and black signature (fixes black signature display issue)
    """
    try:
        print(f"🖊️ Processing signature: {uploaded_file.name}")
        print(f"   Original size: {uploaded_file.size} bytes")
        
        # Store original filename and content type
        original_name = uploaded_file.name
        original_content_type = uploaded_file.content_type
        
        # Read the uploaded file
        uploaded_file.seek(0)
        image = Image.open(uploaded_file)
        
        print(f"   Image mode: {image.mode}")
        print(f"   Image size: {image.size}")
        
        # Convert to RGBA first if needed
        if image.mode != 'RGBA':
            image = image.convert('RGBA')
        
        # Create a white background
        white_bg = Image.new('RGB', image.size, (255, 255, 255))
        
        # Paste the signature onto white background
        # This converts transparent areas to white
        white_bg.paste(image, mask=image.split()[-1])  # Use alpha channel as mask
        
        # Save to BytesIO
        output = io.BytesIO()
        white_bg.save(output, format='PNG', quality=95)
        output.seek(0)
        
        processed_size = output.getbuffer().nbytes
        print(f"   ✓ Processed size: {processed_size} bytes")
        
        # Create new InMemoryUploadedFile with ORIGINAL FILENAME
        processed_file = InMemoryUploadedFile(
            output,
            'ImageField',
            original_name,  # ✅ Keep the original filename
            'image/png',
            processed_size,
            None
        )
        
        # ✅ FIX: Add the 'name' attribute that save_categorized_eligibility_file expects
        processed_file.name = original_name
        
        return processed_file
        
    except Exception as e:
        print(f"   ⚠️ Error processing signature: {e}")
        import traceback
        print(traceback.format_exc())
        # Return original file if processing fails
        uploaded_file.seek(0)
        return uploaded_file

@require_http_methods(["GET"])
def setup_certificate_folders(request):
    """
    Create all required certificate folder structures
    """
    from django.core.files.storage import default_storage
    import os
    
    categories = ['appointive_certificates', 'elective_certificates', 'ids', 'signatures']
    results = {}
    
    for category in categories:
        base_path = f'certification_files/{category}/'
        
        try:
            # Create year/month structure for current year
            from datetime import datetime
            current_year = datetime.now().strftime('%Y')
            current_month = datetime.now().strftime('%m')
            
            full_path = f'{base_path}{current_year}/{current_month}/'
            
            # Create a dummy file to ensure folder exists
            dummy_file = f'{full_path}.gitkeep'
            
            if not default_storage.exists(dummy_file):
                from django.core.files.base import ContentFile
                default_storage.save(dummy_file, ContentFile(b''))
                results[category] = f'✓ Created: {full_path}'
            else:
                results[category] = f'✓ Already exists: {full_path}'
                
        except Exception as e:
            results[category] = f'✗ Error: {str(e)}'
    
    return JsonResponse({
        'success': True,
        'message': 'Certificate folders setup complete',
        'details': results
    }, json_dumps_params={'indent': 2})




#-----------------REQUIREMENTS_MONITORING--------------
@require_http_methods(["GET"])
def user_profile(request):
    return JsonResponse({
        'success': True,
        'user': {
            'full_name': request.user.get_full_name(),
            'email': request.user.email,
            'username': request.user.username,
            'role': request.user.groups.first().name if request.user.groups.exists() else 'User',
            'barangay': request.user.profile.barangay.name if hasattr(request.user, 'profile') else 'Not assigned',
            'member_since': request.user.date_joined.strftime('%B %d, %Y')
        }
    })
@login_required
def requirements_monitoring(request):
    user_profile = request.user.userprofile  
    if user_profile.role == 'dilg staff':
        messages.info(request, 'As DILG Admin, please use the Admin Submissions page.')
        return redirect('admin_submissions')
    barangays = Barangay.objects.all()
    barangay_statuses = {}
    today = date.today()   
    for barangay in barangays:
        submissions = RequirementSubmission.objects.filter(barangay=barangay)
        
        if not submissions.exists():
            # No requirements yet
            barangay_statuses[barangay.id] = {
                'status': 'no_data',
                'color': 'gray',
                'tooltip': 'No requirements assigned'
            }
            continue
        
        # Count different statuses
        total = submissions.count()
        overdue = submissions.filter(
            Q(status__in=['pending', 'in_progress']) & Q(due_date__lt=today)
        ).count()
        pending = submissions.filter(status='pending').count()
        in_progress = submissions.filter(status='in_progress').count()
        accomplished = submissions.filter(status='accomplished').count()
        
        # Determine color based on priority
        if overdue > 0:
            status = 'overdue'
            color = 'red'
            tooltip = f'{overdue} overdue requirement(s)'
        elif pending > 0 or in_progress > 0:
            status = 'in_progress'
            color = 'yellow'
            tooltip = f'{pending} pending, {in_progress} in progress'
        elif accomplished == total:
            status = 'completed'
            color = 'green'
            tooltip = f'All {total} requirements completed!'
        else:
            status = 'partial'
            color = 'blue'
            tooltip = f'{accomplished}/{total} completed'
        
        barangay_statuses[barangay.id] = {
            'status': status,
            'color': color,
            'tooltip': tooltip,
            'counts': {
                'total': total,
                'overdue': overdue,
                'pending': pending,
                'in_progress': in_progress,
                'accomplished': accomplished
            }
        }
    
    context = {
        'barangays': barangays,
        'barangay_statuses': barangay_statuses,
        'user_role': user_profile.role,
        'is_submitter': user_profile.role == 'barangay official',
    }
    return render(request, 'requirements_monitoring.html', context)


import logging
logger = logging.getLogger(__name__)
@login_required
def get_barangay_status(request, barangay_id):

    try:
        barangay = Barangay.objects.get(id=barangay_id)
        submissions = RequirementSubmission.objects.filter(barangay=barangay)
        today = date.today()
        
        if not submissions.exists():
            return JsonResponse({
                'status': 'no_data',
                'color': 'gray',
                'tooltip': f'{barangay.name}: No requirements assigned',
                'counts': {
                    'total': 0,
                    'overdue': 0,
                    'pending': 0,
                    'in_progress': 0,
                    'accomplished': 0,
                    'approved': 0,
                    'rejected': 0
                }
            })
        
        total = submissions.count()

        overdue = submissions.filter(
            status__in=['pending', 'in_progress', 'accomplished'],
            due_date__lt=today
        ).count()
        
        pending = submissions.filter(status='pending').count()
        in_progress = submissions.filter(status='in_progress').count()
        accomplished = submissions.filter(status='accomplished').count()
        approved = submissions.filter(status='approved').count()
        rejected = submissions.filter(status='rejected').count()
        
        if overdue > 0:
            return JsonResponse({
                'status': 'overdue',
                'color': 'red',
                'tooltip': f'{barangay.name}: {overdue} overdue requirement(s) ⚠️',
                'counts': {
                    'total': total,
                    'overdue': overdue,
                    'pending': pending,
                    'in_progress': in_progress,
                    'accomplished': accomplished,
                    'approved': approved,
                    'rejected': rejected
                }
            })
        

        elif approved == total:
            return JsonResponse({
                'status': 'completed',
                'color': 'green',
                'tooltip': f'{barangay.name}: All {total} requirements approved! ✓',
                'counts': {
                    'total': total,
                    'overdue': 0,
                    'pending': 0,
                    'in_progress': 0,
                    'accomplished': 0,
                    'approved': approved,
                    'rejected': rejected
                }
            })

        elif in_progress > 0 or accomplished > 0:
            return JsonResponse({
                'status': 'in_progress',
                'color': 'yellow',
                'tooltip': f'{barangay.name}: {in_progress} in progress, {accomplished} awaiting review',
                'counts': {
                    'total': total,
                    'overdue': 0,
                    'pending': pending,
                    'in_progress': in_progress,
                    'accomplished': accomplished,
                    'approved': approved,
                    'rejected': rejected
                }
            })
        
        elif pending > 0:
            return JsonResponse({
                'status': 'pending',
                'color': 'blue',
                'tooltip': f'{barangay.name}: {pending} pending requirements',
                'counts': {
                    'total': total,
                    'overdue': 0,
                    'pending': pending,
                    'in_progress': 0,
                    'accomplished': 0,
                    'approved': approved,
                    'rejected': rejected
                }
            })

        else:
            return JsonResponse({
                'status': 'partial',
                'color': 'blue',
                'tooltip': f'{barangay.name}: {approved}/{total} approved',
                'counts': {
                    'total': total,
                    'overdue': 0,
                    'pending': pending,
                    'in_progress': in_progress,
                    'accomplished': accomplished,
                    'approved': approved,
                    'rejected': rejected
                }
            })
            
    except Barangay.DoesNotExist:
        return JsonResponse({
            'error': 'Barangay not found',
            'status': 'error',
            'color': 'gray'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'status': 'error',
            'color': 'gray'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_requirements_list(request):
    """API endpoint to list requirements for a barangay"""
    try:
        barangay_id = request.GET.get('barangay_id')
        period = request.GET.get('period', 'weekly')
        week = request.GET.get('week', 1)
        search = request.GET.get('search', '').strip()
        
        if not barangay_id:
            return JsonResponse({'success': False, 'error': 'Barangay ID required'}, status=400)
        
        barangay = get_object_or_404(Barangay, id=barangay_id)
        
        # Get submissions for this barangay and period
        submissions = RequirementSubmission.objects.filter(
            barangay=barangay,
            requirement__period=period,
            requirement__is_active=True
        )
        
        # Filter by week if period is weekly
        if period == 'weekly':
            submissions = submissions.filter(week_number=week)
        
        # Search filter
        if search:
            submissions = submissions.filter(
                Q(requirement__title__icontains=search) |
                Q(requirement__description__icontains=search)
            )
        
        # Prepare response data
        submissions_data = []
        for sub in submissions:
            submissions_data.append({
                'id': sub.id,
                'title': sub.requirement.title,
                'description': sub.requirement.description,
                'status': sub.status,
                'status_display': sub.get_status_display(),
                'due_date': sub.due_date.strftime('%B %d, %Y'),
                'is_overdue': sub.is_overdue,
                'last_update': sub.updated_at.strftime('%B %d, %Y at %I:%M %p'),
                'update_text': sub.update_text,
            })
        
        return JsonResponse({
            'success': True,
            'submissions': submissions_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["GET"])
def api_submission_detail(request, submission_id):
    """API endpoint to get submission details"""
    try:
        # Use select_related and prefetch_related for efficiency
        submission = get_object_or_404(
            RequirementSubmission.objects.select_related(
                'requirement', 'barangay', 'submitted_by'
            ).prefetch_related('attachments'),
            id=submission_id
        )
        
        # Get attachments safely
        attachments = []
        for att in submission.attachments.all():
            try:
                file_name = os.path.basename(att.file.name) if att.file else 'Unknown'
                file_url = att.file.url if att.file else ''
                file_size = att.file_size_kb if hasattr(att, 'file_size_kb') else round(att.file_size / 1024, 2)
                
                attachments.append({
                    'id': att.id,
                    'file_name': file_name,
                    'file_size': file_size,
                    'file_url': file_url,
                    'uploaded_at': att.uploaded_at.strftime('%B %d, %Y at %I:%M %p'),
                })
            except Exception as att_error:
                print(f"Error processing attachment {att.id}: {att_error}")
                continue
        
        data = {
            'id': submission.id,
            'requirement': {
                'title': submission.requirement.title,
                'description': submission.requirement.description,
                'period': submission.requirement.period,
            },
            'barangay': {
                'id': submission.barangay.id,
                'name': submission.barangay.name,
            },
            'status': submission.status,
            'status_display': submission.get_status_display(),
            'due_date': submission.due_date.strftime('%B %d, %Y'),
            'is_overdue': submission.is_overdue,
            'update_text': submission.update_text or '',
            'last_update': submission.updated_at.strftime('%B %d, %Y at %I:%M %p'),
            'attachments': attachments,
        }
        
        return JsonResponse({
            'success': True,
            'submission': data
        })
        
    except RequirementSubmission.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': f'Submission with ID {submission_id} not found'
        }, status=404)
    except Exception as e:
        # Log the full error for debugging
        import traceback
        print(f"=== ERROR in api_submission_detail ===")
        print(f"Submission ID: {submission_id}")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False, 
            'error': f'Internal server error: {str(e)}'
        }, status=500)


@login_required
def debug_submission(request, submission_id):
    try:
        sub = RequirementSubmission.objects.get(id=submission_id)
        return JsonResponse({
            'submission_exists': True,
            'requirement_title': sub.requirement.title,
            'barangay_name': sub.barangay.name,
            'attachment_count': sub.attachments.count(),
            'status': sub.status
        })
    except Exception as e:
        return JsonResponse({'error': str(e)})

@login_required
@require_http_methods(["POST"])
def api_submission_update(request, submission_id):
    """API endpoint to update submission text"""
    try:
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        update_text = request.POST.get('update_text', '').strip()
        
        if not update_text:
            return JsonResponse({'success': False, 'error': 'Update text required'}, status=400)
        
        submission.update_text = update_text
        submission.save()
        
        # Log the update
        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            content_object=submission,
            description=f"Updated requirement: {submission.requirement.title}"
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Update saved successfully',
            'submission': {
                'id': submission.id,
                'update_text': submission.update_text,
                'last_update': submission.updated_at.strftime('%B %d, %Y at %I:%M %p'),
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
def api_attachment_upload(request):
    """
    FIXED: API endpoint to upload file attachments with proper error handling
    """
    try:
        # Debug logging
        print("\n" + "="*60)
        print("📁 FILE UPLOAD REQUEST")
        print("="*60)
        print(f"User: {request.user.username}")
        print(f"POST data: {dict(request.POST)}")
        print(f"FILES: {list(request.FILES.keys())}")
        
        # Get submission_id
        submission_id = request.POST.get('submission_id')
        
        if not submission_id:
            return JsonResponse({
                'success': False, 
                'error': 'Submission ID is required'
            }, status=400)
        
        print(f"Looking for submission: {submission_id}")
        
        # Get submission
        try:
            submission = RequirementSubmission.objects.select_related(
                'requirement', 'barangay'
            ).get(id=submission_id)
            print(f"✓ Found: {submission.requirement.title} - {submission.barangay.name}")
        except RequirementSubmission.DoesNotExist:
            print(f"✗ Submission {submission_id} not found")
            return JsonResponse({
                'success': False, 
                'error': f'Submission {submission_id} not found'
            }, status=404)
        
        # Check for file
        if 'file' not in request.FILES:
            return JsonResponse({
                'success': False, 
                'error': 'No file provided'
            }, status=400)
        
        uploaded_file = request.FILES['file']
        print(f"File: {uploaded_file.name}")
        print(f"Size: {uploaded_file.size} bytes ({round(uploaded_file.size/1024, 2)} KB)")
        print(f"Type: {uploaded_file.content_type}")
        
        # Validate file type
        allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'image/webp']
        if uploaded_file.content_type not in allowed_types:
            return JsonResponse({
                'success': False, 
                'error': f'Invalid file type: {uploaded_file.content_type}. Only images allowed.'
            }, status=400)
        
        # Validate file size (5MB)
        max_size = 5 * 1024 * 1024
        if uploaded_file.size > max_size:
            return JsonResponse({
                'success': False, 
                'error': f'File too large ({round(uploaded_file.size/(1024*1024), 2)}MB). Max: 5MB'
            }, status=400)
        
        # Create attachment - WRAP IN TRY/EXCEPT
        print("Creating RequirementAttachment...")
        try:
            attachment = RequirementAttachment.objects.create(
                submission=submission,
                file=uploaded_file,
                file_type=uploaded_file.content_type,
                file_size=uploaded_file.size,
                uploaded_by=request.user
            )
            print(f"✓ Attachment created: ID {attachment.id}")
        except Exception as create_error:
            print(f"✗ Failed to create attachment: {create_error}")
            print(f"Traceback:\n{traceback.format_exc()}")
            return JsonResponse({
                'success': False,
                'error': f'Failed to save file: {str(create_error)}'
            }, status=500)
        
        # Log the upload (optional)
        try:
            AuditLog.objects.create(
                user=request.user,
                action='CREATE',
                content_object=attachment,
                description=f"Uploaded: {uploaded_file.name} for {submission.requirement.title}"
            )
        except Exception as log_err:
            print(f"⚠️ Audit log failed: {log_err}")
        
        # Prepare response
        try:
            response_data = {
                'success': True,
                'message': 'File uploaded successfully',
                'attachment': {
                    'id': attachment.id,
                    'file_name': os.path.basename(attachment.file.name),
                    'file_size': attachment.file_size_kb,
                    'file_url': attachment.file.url,
                    'uploaded_at': attachment.uploaded_at.strftime('%B %d, %Y at %I:%M %p'),
                }
            }
        except Exception as response_err:
            print(f"⚠️ Error building response: {response_err}")
            # Fallback response
            response_data = {
                'success': True,
                'message': 'File uploaded successfully',
                'attachment': {
                    'id': attachment.id,
                    'file_name': uploaded_file.name,
                    'file_size': round(uploaded_file.size / 1024, 2),
                    'file_url': '',
                    'uploaded_at': timezone.now().strftime('%B %d, %Y at %I:%M %p'),
                }
            }
        
        print(f"✓ SUCCESS")
        print(f"Response: {response_data}")
        print("="*60 + "\n")
        
        return JsonResponse(response_data)
        
    except Exception as e:
        print("\n" + "="*60)
        print("❌ UPLOAD ERROR")
        print("="*60)
        print(f"Error: {str(e)}")
        print(f"Type: {type(e).__name__}")
        print(f"Traceback:\n{traceback.format_exc()}")
        print("="*60 + "\n")
        
        return JsonResponse({
            'success': False, 
            'error': f'Upload failed: {str(e)}'
        }, status=500)



@login_required
@require_http_methods(["POST"])
def api_attachment_delete(request, attachment_id):
    """API endpoint to delete attachment"""
    try:
        attachment = get_object_or_404(RequirementAttachment, id=attachment_id)
        
        # Check if user has permission to delete
        if attachment.uploaded_by != request.user and not request.user.is_staff:
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        # Log before deletion
        AuditLog.objects.create(
            user=request.user,
            action='DELETE',
            description=f"Deleted attachment: {os.path.basename(attachment.file.name)}"
        )
        
        attachment.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'File removed successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
def api_submission_submit(request, submission_id):
    """
    API endpoint to submit requirement to admin
    When barangay official submits, DILG admin automatically sees it
    """
    try:
        # 🔒 Access Control - Only Barangay Officials can submit
        user_profile = request.user.userprofile
        if user_profile.role not in ['barangay official', 'municipal officer']:
            return JsonResponse({
                'success': False,
                'error': 'Only Barangay Officials can submit requirements'
            }, status=403)
        
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        # Get update_text from POST
        update_text = request.POST.get('update_text', '').strip()
        
        if update_text:
            submission.update_text = update_text
            submission.save()
        
        # Validate required fields
        if not submission.update_text:
            return JsonResponse({
                'success': False, 
                'error': 'Please add update details before submitting'
            }, status=400)
        
        if not submission.attachments.exists():
            return JsonResponse({
                'success': False,
                'error': 'Please upload at least one image before submitting'
            }, status=400)
        
        # ✅ SUBMIT TO ADMIN - Change status to 'accomplished'
        # This makes it visible to DILG admin immediately
        submission.status = 'accomplished'
        submission.submitted_by = request.user
        submission.submitted_at = timezone.now()
        submission.save()
        
        # 📧 Optional: Send email notification to DILG admin
        # (You can add this if needed)
        try:
            from django.core.mail import send_mail
            send_mail(
                subject=f'New Requirement Submitted: {submission.requirement.title}',
                message=f'{submission.barangay.name} submitted: {submission.requirement.title}',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[settings.EMAIL_HOST_USER],  # DILG admin email
                fail_silently=True,
            )
        except:
            pass  # Don't fail submission if email fails
        
        # Log the submission
        try:
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                content_object=submission,
                description=f"Submitted requirement to DILG Admin: {submission.requirement.title} from {submission.barangay.name}"
            )
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': '✅ Successfully submitted to DILG Admin!',
            'submission': {
                'id': submission.id,
                'status': submission.status,
                'status_display': submission.get_status_display(),
                'submitted_at': submission.submitted_at.strftime('%B %d, %Y at %I:%M %p')
            }
        })
        
    except Exception as e:
        import traceback
        print(f"=== SUBMIT ERROR ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=500)



@login_required
@require_http_methods(["POST"])
def api_submission_delete(request, submission_id):
    """API endpoint to delete submission"""
    try:
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        # Check if user has permission
        if not request.user.is_staff:
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        # Log before deletion
        AuditLog.objects.create(
            user=request.user,
            action='DELETE',
            description=f"Deleted submission: {submission.requirement.title} - {submission.barangay.name}"
        )
        
        submission.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Requirement deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def requirements_monitoring(request):
    """
    Requirements submission page - BARANGAY OFFICIALS ONLY
    This is where barangay officials submit their requirements
    """
    user_profile = request.user.userprofile
    
    # 🔒 STRICT ACCESS CONTROL - Only Barangay Officials
    if user_profile.role == 'dilg staff':
        messages.warning(request, '⚠️ DILG Admin should use the Admin Submissions page to review submissions.')
        return redirect('admin_submissions_page')  # Redirect to admin page
    
    if user_profile.role not in ['barangay official', 'municipal officer']:
        messages.error(request, '🚫 Access Denied: This page is only for Barangay Officials.')
        return redirect('dashboard')
    
    # Get barangays for barangay officials
    barangays = Barangay.objects.all().order_by('id') 
    
    context = {
        'barangays': barangays,
        'user_role': user_profile.role,
        'page_title': 'Submit Requirements',
        'is_submitter': True,  # Flag to show this is submission page
    }
    return render(request, 'requirements_monitoring.html', context)


@login_required
@require_http_methods(["GET"])
def get_requirements_list(request):
    """AJAX endpoint to get requirements list for a barangay"""
    try:
        barangay_id = request.GET.get('barangay_id')
        period = request.GET.get('period', 'weekly')
        week = request.GET.get('week', 1)
        search = request.GET.get('search', '')
        
        if not barangay_id:
            return JsonResponse({
                'success': False,
                'error': 'Barangay ID is required'
            }, status=400)
        
        barangay = get_object_or_404(Barangay, id=barangay_id)
        
        # Get current year and week
        current_year = timezone.now().year
        current_week = int(week)
        
        # Get requirements for this period
        requirements = Requirement.objects.filter(
            Q(period=period) & Q(is_active=True)
        ).filter(
            Q(applicable_barangays=barangay) | Q(applicable_barangays__isnull=True)
        )
        
        # Apply search filter
        if search:
            requirements = requirements.filter(
                Q(title__icontains=search) | Q(description__icontains=search)
            )
        
        # Get or create submissions for these requirements
        submissions_data = []
        for req in requirements:
            # Get or create submission for this week/period
            submission, created = RequirementSubmission.objects.get_or_create(
                requirement=req,
                barangay=barangay,
                week_number=current_week if period == 'weekly' else None,
                year=current_year,
                defaults={
                    'due_date': calculate_due_date(period, current_week, current_year),
                    'status': 'pending'
                }
            )
            
            submissions_data.append({
                'id': submission.id,
                'requirement_id': req.id,
                'title': req.title,
                'description': req.description,
                'status': submission.status,
                'status_display': submission.get_status_display(),
                'due_date': submission.due_date.strftime('%B %d, %Y'),
                'last_update': submission.updated_at.strftime('%B %d, %Y'),
                'is_overdue': submission.is_overdue,
                'has_attachments': submission.attachments.exists(),
                'attachment_count': submission.attachments.count(),
            })
        
        return JsonResponse({
            'success': True,
            'submissions': submissions_data,
            'barangay_name': barangay.name,
            'period': period,
            'week': current_week
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["GET"])
def get_submission_detail(request, submission_id):
    """Get detailed information about a specific submission"""
    try:
        submission = get_object_or_404(
            RequirementSubmission.objects.select_related(
                'requirement', 'barangay', 'submitted_by', 'reviewed_by'
            ).prefetch_related('attachments'),
            id=submission_id
        )
        
        # Get attachments
        attachments = []
        for attachment in submission.attachments.all():
            attachments.append({
                'id': attachment.id,
                'file_url': attachment.file.url,
                'file_name': attachment.file.name.split('/')[-1],
                'file_size': attachment.file_size_kb,
                'file_type': attachment.file_type,
                'uploaded_at': attachment.uploaded_at.strftime('%B %d, %Y %I:%M %p')
            })
        
        data = {
            'id': submission.id,
            'requirement': {
                'title': submission.requirement.title,
                'description': submission.requirement.description,
                'period': submission.requirement.get_period_display()
            },
            'barangay': {
                'name': submission.barangay.name,
                'code': submission.barangay.code
            },
            'status': submission.status,
            'status_display': submission.get_status_display(),
            'due_date': submission.due_date.strftime('%B %d, %Y'),
            'week_number': submission.week_number,
            'year': submission.year,
            'update_text': submission.update_text or '',
            'is_overdue': submission.is_overdue,
            'submitted_by': submission.submitted_by.get_full_name() if submission.submitted_by else None,
            'submitted_at': submission.submitted_at.strftime('%B %d, %Y %I:%M %p') if submission.submitted_at else None,
            'reviewed_by': submission.reviewed_by.get_full_name() if submission.reviewed_by else None,
            'reviewed_at': submission.reviewed_at.strftime('%B %d, %Y %I:%M %p') if submission.reviewed_at else None,
            'review_notes': submission.review_notes or '',
            'attachments': attachments,
            'last_update': submission.updated_at.strftime('%B %d, %Y %I:%M %p')
        }
        
        return JsonResponse({
            'success': True,
            'submission': data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def update_submission(request, submission_id):
    """Update a requirement submission"""
    try:
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        data = json.loads(request.body)
        update_text = data.get('update_text', '').strip()
        
        if not update_text:
            return JsonResponse({
                'success': False,
                'error': 'Update text is required'
            }, status=400)
        
        # Update submission
        submission.update_text = update_text
        submission.status = 'in_progress'
        submission.updated_at = timezone.now()
        submission.save()
        
        # Log the update
        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            content_object=submission,
            description=f"Updated requirement submission: {submission.requirement.title}"
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Submission updated successfully',
            'last_update': submission.updated_at.strftime('%B %d, %Y %I:%M %p')
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def upload_attachment(request, submission_id):
    """Upload file attachments for a submission"""
    try:
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        # Get uploaded files
        files = request.FILES.getlist('files')
        
        if not files:
            return JsonResponse({
                'success': False,
                'error': 'No files uploaded'
            }, status=400)
        
        uploaded_files = []
        
        for file in files:
            # Validate file size (max 5MB)
            if file.size > 5 * 1024 * 1024:
                return JsonResponse({
                    'success': False,
                    'error': f'File {file.name} is too large (max 5MB)'
                }, status=400)
            
            # Create attachment
            attachment = RequirementAttachment.objects.create(
                submission=submission,
                file=file,
                file_type=file.content_type,
                file_size=file.size,
                uploaded_by=request.user
            )
            
            uploaded_files.append({
                'id': attachment.id,
                'file_name': file.name,
                'file_size': attachment.file_size_kb,
                'file_url': attachment.file.url
            })
        
        # Log the upload
        AuditLog.objects.create(
            user=request.user,
            action='CREATE',
            content_object=submission,
            description=f"Uploaded {len(files)} file(s) for requirement: {submission.requirement.title}"
        )
        
        return JsonResponse({
            'success': True,
            'message': f'{len(files)} file(s) uploaded successfully',
            'files': uploaded_files
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["DELETE"])
def delete_attachment(request, attachment_id):
    """Delete a file attachment"""
    try:
        attachment = get_object_or_404(RequirementAttachment, id=attachment_id)
        submission = attachment.submission
        
        # Delete the file
        attachment.delete()
        
        # Log the deletion
        AuditLog.objects.create(
            user=request.user,
            action='DELETE',
            content_object=submission,
            description=f"Deleted attachment for requirement: {submission.requirement.title}"
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Attachment deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def submit_to_admin(request, submission_id):
    """Submit requirement to admin for review"""
    try:
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        # Validate that there's content
        if not submission.update_text:
            return JsonResponse({
                'success': False,
                'error': 'Please add update details before submitting'
            }, status=400)
        
        # Update submission status
        submission.submit(request.user)
        
        return JsonResponse({
            'success': True,
            'message': 'Requirement submitted to admin successfully',
            'new_status': submission.status,
            'submitted_at': submission.submitted_at.strftime('%B %d, %Y %I:%M %p')
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["DELETE"])
def delete_submission(request, submission_id):
    """Delete a requirement submission"""
    try:
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        # Only allow deletion if status is pending
        if submission.status not in ['pending', 'in_progress']:
            return JsonResponse({
                'success': False,
                'error': 'Cannot delete submitted or reviewed requirements'
            }, status=400)
        
        requirement_title = submission.requirement.title
        
        # Log before deletion
        AuditLog.objects.create(
            user=request.user,
            action='DELETE',
            description=f"Deleted requirement submission: {requirement_title}"
        )
        
        # Delete submission (attachments will be deleted via cascade)
        submission.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Requirement "{requirement_title}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# Helper function
def calculate_due_date(period, week_number, year):
    """Calculate due date based on period and week number"""
    if period == 'weekly':
        # Calculate the last day of the specified week
        jan_1 = datetime(year, 1, 1)
        days_to_add = (week_number - 1) * 7 + (6 - jan_1.weekday())
        return (jan_1 + timedelta(days=days_to_add)).date()
    
    elif period == 'monthly':
        # Last day of current month
        if timezone.now().month == 12:
            return datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            return datetime(year, timezone.now().month + 1, 1).date() - timedelta(days=1)
    
    elif period == 'quarterly':
        # End of current quarter
        current_quarter = (timezone.now().month - 1) // 3
        quarter_end_month = (current_quarter + 1) * 3
        if quarter_end_month == 12:
            return datetime(year, 12, 31).date()
        else:
            return datetime(year, quarter_end_month + 1, 1).date() - timedelta(days=1)
    
    elif period == 'semestral':
        # End of semester (June 30 or December 31)
        if timezone.now().month <= 6:
            return datetime(year, 6, 30).date()
        else:
            return datetime(year, 12, 31).date()
    
    elif period == 'annually':
        # End of year
        return datetime(year, 12, 31).date()
    
    return timezone.now().date()


#DILG ADMIN
@login_required
def admin_submissions_page(request):
    """
    Admin review page - DILG STAFF ONLY
    This is where DILG admin reviews all submissions from all barangays
    """
    user_profile = request.user.userprofile
    
    # 🔒 STRICT ACCESS CONTROL - Only DILG Staff
    if user_profile.role == 'barangay official':
        messages.error(request, '🚫 Access Denied: Barangay Officials cannot access the admin review page.')
        return redirect('requirements_monitoring')  # Redirect to submission page
    
    if user_profile.role != 'dilg staff':
        messages.error(request, '🚫 Access Denied: This page is only accessible to DILG Admin.')
        return redirect('dashboard')
    
    barangays = Barangay.objects.all().order_by('name')
    
    context = {
        'barangays': barangays,
        'page_title': 'Review Submissions',
        'is_admin_view': True,  # Flag to show this is admin page
    }
    return render(request, 'admin_submissions.html', context)


@login_required
@require_http_methods(["GET"])
def api_admin_submissions_list(request):
    """
    API endpoint to get all submissions for DILG admin review
    ONLY accessible by DILG Staff
    """
    try:
        # STRICT ACCESS CONTROL
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized: Only DILG Admin can view submissions'
            }, status=403)
        
        # Get filter parameters
        barangay_id = request.GET.get('barangay_id', '')
        status_filter = request.GET.get('status', '')
        period_filter = request.GET.get('period', '')
        search_query = request.GET.get('search', '').strip()
        
        # Base query - get ALL submissions from ALL barangays
        submissions = RequirementSubmission.objects.select_related(
            'requirement', 'barangay', 'submitted_by', 'reviewed_by'
        ).prefetch_related('attachments')
        
        # Apply filters
        if barangay_id:
            submissions = submissions.filter(barangay_id=barangay_id)
        
        if status_filter:
            submissions = submissions.filter(status=status_filter)
        else:
            # Default: show only submitted (accomplished) items awaiting review
            submissions = submissions.filter(status='accomplished')
        
        if period_filter:
            submissions = submissions.filter(requirement__period=period_filter)
        
        if search_query:
            submissions = submissions.filter(
                Q(requirement__title__icontains=search_query) |
                Q(requirement__description__icontains=search_query) |
                Q(barangay__name__icontains=search_query) |
                Q(update_text__icontains=search_query)
            )
        
        # Get statistics for all submissions
        all_submissions = RequirementSubmission.objects.all()
        
        stats = {
            'submitted': all_submissions.filter(status='accomplished').count(),
            'approved': all_submissions.filter(status='approved').count(),
            'rejected': all_submissions.filter(status='rejected').count(),
            'overdue': all_submissions.filter(
                due_date__lt=timezone.now().date(),
                status__in=['pending', 'in_progress', 'accomplished']
            ).count()
        }
        
        # Prepare submissions data
        submissions_data = []
        for sub in submissions.order_by('-submitted_at', '-updated_at'):
            # Get attachments
            attachments = []
            for att in sub.attachments.all():
                try:
                    attachments.append({
                        'id': att.id,
                        'file_name': os.path.basename(att.file.name) if att.file else 'Unknown',
                        'file_url': att.file.url if att.file else '',
                        'file_size': att.file_size_kb if hasattr(att, 'file_size_kb') else 
                                   round(att.file_size / 1024, 2) if hasattr(att, 'file_size') else 0,
                    })
                except:
                    continue
            
            submissions_data.append({
                'id': sub.id,
                'title': sub.requirement.title,
                'description': sub.requirement.description,
                'barangay_name': sub.barangay.name,
                'barangay_id': sub.barangay.id,
                'status': sub.status,
                'status_display': sub.get_status_display(),
                'period': sub.requirement.get_period_display(),
                'due_date': sub.due_date.strftime('%B %d, %Y'),
                'is_overdue': sub.is_overdue,
                'update_text': sub.update_text or '',
                'submitted_at': sub.submitted_at.strftime('%B %d, %Y') if sub.submitted_at else None,
                'submitted_by': sub.submitted_by.get_full_name() if sub.submitted_by else None,
                'reviewed_at': sub.reviewed_at.strftime('%B %d, %Y') if sub.reviewed_at else None,
                'reviewed_by': sub.reviewed_by.get_full_name() if sub.reviewed_by else None,
                'review_notes': sub.review_notes or '',
                'attachments': attachments,
            })
        
        return JsonResponse({
            'success': True,
            'submissions': submissions_data,
            'stats': stats
        })
        
    except Exception as e:
        import traceback
        print(f"=== API ADMIN SUBMISSIONS ERROR ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_admin_review_submission(request, submission_id):
    """
    FIXED: API endpoint for DILG admin to approve/reject submissions
    ONLY accessible by DILG Staff
    """
    try:
        # STRICT ACCESS CONTROL
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized: Only DILG Admin can review submissions'
            }, status=403)
        
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        data = json.loads(request.body)
        action = data.get('action')  # 'approved' or 'rejected'
        review_notes = data.get('review_notes', '').strip()
        
        if action not in ['approved', 'rejected']:
            return JsonResponse({
                'success': False,
                'error': 'Invalid action. Must be "approved" or "rejected"'
            }, status=400)
        
        # Update submission
        submission.status = action
        submission.reviewed_by = request.user
        submission.reviewed_at = timezone.now()
        submission.review_notes = review_notes
        submission.save()
        
        # Log the review
        try:
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                content_object=submission,
                description=f"DILG Admin {action.upper()} submission: {submission.requirement.title} from {submission.barangay.name}"
            )
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Submission {action} successfully',
            'submission': {
                'id': submission.id,
                'status': submission.status,
                'status_display': submission.get_status_display(),
                'reviewed_by': request.user.get_full_name() or request.user.username,
                'reviewed_at': submission.reviewed_at.strftime('%B %d, %Y %I:%M %p')
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        import traceback
        print(f"=== REVIEW ERROR ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def check_page_access(request, allowed_roles):
    """
    Reusable function to check if user has access to a page
    Returns: (has_access: bool, redirect_url: str or None)
    """
    if not request.user.is_authenticated:
        return False, 'login_page'
    
    try:
        user_profile = request.user.userprofile
        if user_profile.role in allowed_roles:
            return True, None
        
        # Determine where to redirect based on role
        if user_profile.role == 'dilg staff':
            return False, 'admin_submissions_page'
        elif user_profile.role == 'barangay official':
            return False, 'requirements_monitoring'
        else:
            return False, 'dashboard'
    except:
        return False, 'dashboard'
    

@login_required
@require_http_methods(["POST"])
def api_create_requirement(request):
    """
    API endpoint for DILG Admin to create new requirements
    ONLY accessible by DILG Staff
    """
    try:
        # STRICT ACCESS CONTROL
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized: Only DILG Admin can create requirements'
            }, status=403)
        
        data = json.loads(request.body)
        
        title = data.get('title', '').strip()
        description = data.get('description', '').strip()
        period = data.get('period', '').strip()
        barangay_ids = data.get('barangay_ids', [])  # Empty = all barangays
        
        # Validation
        if not title:
            return JsonResponse({'success': False, 'error': 'Title is required'}, status=400)
        
        if not description:
            return JsonResponse({'success': False, 'error': 'Description is required'}, status=400)
        
        if not period:
            return JsonResponse({'success': False, 'error': 'Period is required'}, status=400)
        
        valid_periods = ['weekly', 'monthly', 'quarterly', 'semestral', 'annually']
        if period not in valid_periods:
            return JsonResponse({
                'success': False,
                'error': f'Invalid period. Must be one of: {valid_periods}'
            }, status=400)
        
        # Create requirement
        requirement = Requirement.objects.create(
            title=title,
            description=description,
            period=period,
            created_by=request.user,
            is_active=True
        )
        
        # Assign to specific barangays if provided, otherwise applies to all
        if barangay_ids:
            target_barangays = Barangay.objects.filter(id__in=barangay_ids)
            requirement.applicable_barangays.set(target_barangays)
        else:
            # If no specific barangays, get all
            target_barangays = Barangay.objects.all()
        
        # 🔥 FIX: AUTO-CREATE SUBMISSIONS FOR ALL BARANGAYS
        current_year = timezone.now().year
        submissions_created = 0
        
        for barangay in target_barangays:
            if period == 'weekly':
                # Create submissions for the next 4 weeks
                for week_num in range(1, 5):
                    RequirementSubmission.objects.create(
                        requirement=requirement,
                        barangay=barangay,
                        week_number=week_num,
                        year=current_year,
                        due_date=calculate_due_date(period, week_num, current_year),
                        status='pending'
                    )
                    submissions_created += 1
            else:
                # For monthly, quarterly, semestral, annually - create one submission
                RequirementSubmission.objects.create(
                    requirement=requirement,
                    barangay=barangay,
                    week_number=None,
                    year=current_year,
                    due_date=calculate_due_date(period, 1, current_year),
                    status='pending'
                )
                submissions_created += 1
        
        print(f"✅ Created {submissions_created} submissions for requirement: {title}")
        
        # Log the creation
        try:
            AuditLog.objects.create(
                user=request.user,
                action='CREATE',
                content_object=requirement,
                description=f"DILG Admin created new requirement: {title} with {submissions_created} submissions"
            )
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Requirement created successfully! {submissions_created} submissions created for barangays.',
            'requirement': {
                'id': requirement.id,
                'title': requirement.title,
                'description': requirement.description,
                'period': requirement.period,
                'period_display': requirement.get_period_display(),
                'submissions_created': submissions_created,
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        import traceback
        print(f"=== CREATE REQUIREMENT ERROR ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def api_edit_requirement(request, requirement_id):
    """
    API endpoint for DILG Admin to edit existing requirements
    ONLY accessible by DILG Staff
    """
    try:
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized: Only DILG Admin can edit requirements'
            }, status=403)
        
        requirement = get_object_or_404(Requirement, id=requirement_id)
        
        data = json.loads(request.body)
        
        title = data.get('title', '').strip()
        description = data.get('description', '').strip()
        period = data.get('period', '').strip()
        is_active = data.get('is_active', True)
        
        if title:
            requirement.title = title
        if description:
            requirement.description = description
        if period:
            requirement.period = period
        
        requirement.is_active = is_active
        requirement.save()
        
        # Log the update
        try:
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                content_object=requirement,
                description=f"DILG Admin updated requirement: {requirement.title}"
            )
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': 'Requirement updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



@login_required
@require_http_methods(["DELETE"])
def api_delete_requirement(request, requirement_id):
    """
    API endpoint for DILG Admin to delete requirements
    ONLY accessible by DILG Staff
    """
    try:
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized: Only DILG Admin can delete requirements'
            }, status=403)
        
        requirement = get_object_or_404(Requirement, id=requirement_id)
        title = requirement.title
        
        # Log before deletion
        try:
            AuditLog.objects.create(
                user=request.user,
                action='DELETE',
                description=f"DILG Admin deleted requirement: {title}"
            )
        except:
            pass
        
        requirement.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Requirement "{title}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



@login_required
@require_http_methods(["GET"])
def api_all_requirements(request):
    """
    API endpoint to get all requirements (for DILG Admin management)
    ONLY accessible by DILG Staff
    """
    try:
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized'
            }, status=403)
        
        requirements = Requirement.objects.all().order_by('-created_at')
        
        requirements_data = []
        for req in requirements:
            applicable_barangays = list(req.applicable_barangays.values_list('name', flat=True))
            
            requirements_data.append({
                'id': req.id,
                'title': req.title,
                'description': req.description,
                'period': req.period,
                'period_display': req.get_period_display(),
                'is_active': req.is_active,
                'applicable_barangays': applicable_barangays if applicable_barangays else ['All Barangays'],
                'created_at': req.created_at.strftime('%B %d, %Y'),
                'created_by': req.created_by.get_full_name() if req.created_by else 'System',
            })
        
        return JsonResponse({
            'success': True,
            'requirements': requirements_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    



#-------------------ANNOUNCEMENTS AND NOTIFICATIONS------------
@login_required
def delete_announcement(request, announcement_id):
    """Delete an announcement"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Only POST requests are allowed'
        }, status=405)
    
    try:
        print(f"Attempting to delete announcement with ID: {announcement_id}")
        
        # Get the announcement
        announcement = Announcement.objects.get(id=announcement_id)
        print(f"Found announcement: {announcement.title}")
        
        # Store info before deletion
        title = announcement.title
        
        # Delete the announcement
        announcement.delete()
        print(f"Successfully deleted announcement: {title}")
        
        return JsonResponse({
            'success': True,
            'message': f'Announcement "{title}" deleted successfully'
        })
        
    except Announcement.DoesNotExist:
        print(f"Announcement with ID {announcement_id} does not exist")
        return JsonResponse({
            'success': False,
            'error': 'Announcement not found'
        }, status=404)
        
    except Exception as e:
        # Log the full error for debugging
        error_details = traceback.format_exc()
        print(f"Error deleting announcement {announcement_id}:")
        print(error_details)
        
        return JsonResponse({
            'success': False,
            'error': f'Server error: {str(e)}'
        }, status=500)

@login_required
@require_POST
def update_announcement(request, announcement_id):
    """Update an existing announcement"""
    try:
        announcement = Announcement.objects.get(id=announcement_id)
        data = json.loads(request.body)
        
        # Update fields
        announcement.title = data.get('title', announcement.title)
        announcement.content = data.get('content', announcement.content)
        announcement.priority = data.get('priority', announcement.priority)
        announcement.save()
        
        notifications_sent = 0
        send_notification = data.get('send_notification', False)
        
        # Send update notifications if enabled
        if send_notification:
            barangay_users = User.objects.filter(
                userprofile__role='barangay official',
                is_active=True
            ).distinct()
            
            notification_list = []
            for user in barangay_users:
                notification_list.append(
                    Notification(
                        user=user,
                        title=f"📢 Updated: {announcement.title}",
                        message=f"An announcement has been updated: {announcement.content[:100]}...",
                        notification_type='info',
                        created_at=timezone.now()
                    )
                )
            
            Notification.objects.bulk_create(notification_list)
            notifications_sent = len(notification_list)
        
        return JsonResponse({
            'success': True,
            'announcement': {
                'id': announcement.id,
                'title': announcement.title,
                'content': announcement.content,
                'priority': announcement.priority
            },
            'notifications_sent': notifications_sent
        })
        
    except Announcement.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Announcement not found'
        }, status=404)
    except Exception as e:
        import traceback
        print(f"=== UPDATE ANNOUNCEMENT ERROR ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
@login_required
@require_http_methods(["GET"])
def get_announcements(request):
    """Get all announcements"""
    try:
        announcements = Announcement.objects.all().select_related('posted_by').order_by('-posted_at')
        
        announcements_list = []
        for announcement in announcements:
            announcements_list.append({
                'id': announcement.id,
                'title': announcement.title,
                'content': announcement.content,
                'priority': announcement.priority,
                'priority_display': announcement.get_priority_display(),
                'posted_by': announcement.posted_by.get_full_name() or announcement.posted_by.username,
                'posted_at': announcement.posted_at.strftime('%B %d, %Y'),
                'views': announcement.views,
                'sent_to_barangays': 33  
            })
        
        return JsonResponse({
            'success': True,
            'announcements': announcements_list
        })
        
    except Exception as e:
        import traceback
        print(f"=== GET ANNOUNCEMENTS ERROR ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
@login_required
def debug_users(request):
    """Debug view to see all users and their roles"""
    users = User.objects.all()
    
    user_list = []
    for user in users:
        try:
            role = user.userprofile.role if hasattr(user, 'userprofile') else 'No profile'
        except:
            role = 'Error getting role'
        
        user_list.append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'is_active': user.is_active,
            'is_superuser': user.is_superuser,
            'role': role
        })
    
    return JsonResponse({
        'total_users': users.count(),
        'users': user_list
    }, json_dumps_params={'indent': 2})

#-----------------END OF ANNOUNCEMENTS AND NOTIFICATIONS--------

#---------------NOTIFICATIONS----------------------
@login_required
@require_http_methods(["GET"])
def get_notifications(request):
    """Get all notifications for the current user"""
    try:
        notifications = Notification.objects.filter(
            user=request.user
        ).select_related('submission', 'submission__requirement', 'submission__barangay').order_by('-created_at')
        
        unread_count = notifications.filter(is_read=False).count()
        
        notification_list = []
        for notif in notifications[:50]:
            notification_data = {
                'id': notif.id,
                'title': notif.title,
                'message': notif.message,
                'type': notif.notification_type,
                'is_read': notif.is_read,
                'created_at': notif.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'time_ago': get_time_ago(notif.created_at),
            }
            
            if notif.submission:
                notification_data['submission'] = {
                    'id': notif.submission.id,
                    'requirement_name': notif.submission.requirement.title,
                    'barangay_name': notif.submission.barangay.name,
                    'status': notif.submission.status,
                    'due_date': notif.submission.due_date.strftime('%Y-%m-%d'),
                }
            
            notification_list.append(notification_data)
        
        return JsonResponse({
            'success': True,
            'notifications': notification_list,
            'unread_count': unread_count,
            'total_count': notifications.count()
        })
        
    except Exception as e:
        print(f"Error in get_notifications: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_http_methods(["POST"])
def mark_notification_read(request, notification_id):
    """Mark a single notification as read"""
    try:
        notification = Notification.objects.get(
            id=notification_id,
            user=request.user
        )
        notification.is_read = True
        notification.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Notification marked as read'
        })
        
    except Notification.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Notification not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def mark_all_notifications_read(request):
    """Mark all notifications as read for current user"""
    try:
        updated = Notification.objects.filter(
            user=request.user,
            is_read=False
        ).update(is_read=True)
        
        return JsonResponse({
            'success': True,
            'message': f'{updated} notifications marked as read'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_http_methods(["GET"])
def get_unread_count(request):
    """Get count of unread notifications"""
    try:
        count = Notification.objects.filter(
            user=request.user,
            is_read=False
        ).count()
        
        return JsonResponse({
            'success': True,
            'unread_count': count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
@login_required
@require_POST
def create_announcement(request):
    """Create a new announcement and send notifications to barangays"""
    try:
        data = json.loads(request.body)
        
        title = data.get('title', '').strip()
        content = data.get('content', '').strip()
        priority = data.get('priority', 'medium')
        send_notification = data.get('send_notification', False)
        
        # Validation
        if not title or not content:
            return JsonResponse({
                'success': False,
                'error': 'Title and content are required'
            }, status=400)
        
        # Create announcement
        announcement = Announcement.objects.create(
            title=title,
            content=content,
            priority=priority,
            posted_by=request.user,
            posted_at=timezone.now()
        )
        
        notifications_sent = 0
        
        # Send notifications to all barangay users if enabled
        if send_notification:
            # FIX: Get ALL active users (or all users with 'barangay official' role)
            # Option 1: Send to all active users
            barangay_users = User.objects.filter(
                is_active=True,
                is_superuser=False  # Exclude admin from notifications
            ).exclude(
                id=request.user.id  # Exclude the person posting
            ).distinct()
            
            # Debug: Print how many users we found
            print(f"📊 Found {barangay_users.count()} users to notify")
            for user in barangay_users:
                try:
                    role = user.userprofile.role if hasattr(user, 'userprofile') else 'No role'
                    print(f"   - {user.username}: {role}")
                except:
                    print(f"   - {user.username}: No profile")
            
            # Create notification for each user
            notification_list = []
            for user in barangay_users:
                notification_list.append(
                    Notification(
                        user=user,
                        title=f"📢 New Announcement: {title}",
                        message=f"{content[:100]}..." if len(content) > 100 else content,
                        notification_type='info',
                        announcement=announcement,
                        created_at=timezone.now()
                    )
                )
            
            # Bulk create notifications for efficiency
            if notification_list:
                Notification.objects.bulk_create(notification_list)
                notifications_sent = len(notification_list)
                print(f"✅ Created {notifications_sent} notifications")
        
        return JsonResponse({
            'success': True,
            'announcement': {
                'id': announcement.id,
                'title': announcement.title,
                'content': announcement.content,
                'priority': announcement.priority,
                'posted_at': announcement.posted_at.strftime('%B %d, %Y'),
                'posted_by': announcement.posted_by.get_full_name() or announcement.posted_by.username
            },
            'notifications_sent': notifications_sent
        })
        
    except Exception as e:
        import traceback
        print(f"=== CREATE ANNOUNCEMENT ERROR ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

    
@login_required
@require_http_methods(["POST"])
def submit_requirement_with_notification(request, submission_id):
    """Handle requirement submission and create notifications"""
    try:
        submission = RequirementSubmission.objects.get(
            id=submission_id,
            barangay__user=request.user
        )
        
        # Update submission status
        submission.status = 'accomplished'
        submission.submitted_at = timezone.now()
        submission.save()
        
        # Create notification for admins
        create_admin_notification(
            title="New Submission for Review",
            message=f"{submission.barangay.name} submitted {submission.requirement.title}",
            notification_type='info',
            submission=submission
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Requirement submitted successfully',
            'submission': {
                'id': submission.id,
                'status': submission.status,
                'status_display': submission.get_status_display(),
                'submitted_at': submission.submitted_at.strftime('%B %d, %Y')
            }
        })
        
    except RequirementSubmission.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Submission not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_http_methods(["POST"])
def approve_submission_with_notification(request, submission_id):
    """Approve submission and notify submitter"""
    try:
        # Check if user is admin
        if not (request.user.groups.filter(name='Admin').exists() or request.user.is_superuser):
            return JsonResponse({
                'success': False,
                'error': 'Permission denied'
            }, status=403)
        
        submission = RequirementSubmission.objects.get(id=submission_id)
        
        # Get review notes if any
        import json
        data = json.loads(request.body) if request.body else {}
        review_notes = data.get('review_notes', '')
        
        # Update submission
        submission.status = 'approved'
        submission.reviewed_by = request.user
        submission.reviewed_at = timezone.now()
        submission.review_notes = review_notes
        submission.save()
        
        # Create notification for submitter
        message = f"Your submission for {submission.requirement.title} has been approved"
        if review_notes:
            message += f". Admin notes: {review_notes}"
        
        create_notification(
            user=submission.barangay.user,
            title="Submission Approved",
            message=message,
            notification_type='completed',
            submission=submission
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Submission approved successfully'
        })
        
    except RequirementSubmission.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Submission not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_http_methods(["POST"])
def reject_submission_with_notification(request, submission_id):
    """Reject submission and notify submitter"""
    try:
        # Check if user is admin
        if not (request.user.groups.filter(name='Admin').exists() or request.user.is_superuser):
            return JsonResponse({
                'success': False,
                'error': 'Permission denied'
            }, status=403)
        
        submission = RequirementSubmission.objects.get(id=submission_id)
        
        # Get review notes
        import json
        data = json.loads(request.body) if request.body else {}
        review_notes = data.get('review_notes', '')
        
        # Update submission
        submission.status = 'rejected'
        submission.reviewed_by = request.user
        submission.reviewed_at = timezone.now()
        submission.review_notes = review_notes
        submission.save()
        
        # Create notification for submitter
        message = f"Your submission for {submission.requirement.title} needs revision"
        if review_notes:
            message += f". Admin notes: {review_notes}"
        
        create_notification(
            user=submission.barangay.user,
            title="Submission Needs Revision",
            message=message,
            notification_type='overdue',
            submission=submission
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Submission rejected successfully'
        })
        
    except RequirementSubmission.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Submission not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


def get_time_ago(datetime_obj):
    """Convert datetime to 'time ago' format"""
    from django.utils import timezone
    now = timezone.now()
    diff = now - datetime_obj
    
    seconds = diff.total_seconds()
    
    if seconds < 60:
        return "Just now"
    elif seconds < 3600:
        minutes = int(seconds / 60)
        return f"{minutes} minute{'s' if minutes > 1 else ''} ago"
    elif seconds < 86400:
        hours = int(seconds / 3600)
        return f"{hours} hour{'s' if hours > 1 else ''} ago"
    elif seconds < 604800:
        days = int(seconds / 86400)
        return f"{days} day{'s' if days > 1 else ''} ago"
    else:
        weeks = int(seconds / 604800)
        return f"{weeks} week{'s' if weeks > 1 else ''} ago"

def create_notification(user, title, message, notification_type, submission=None):
    """Helper function to create notifications"""
    try:
        notification = Notification.objects.create(
            user=user,
            title=title,
            message=message,
            notification_type=notification_type,
            submission=submission
        )
        return notification
    except Exception as e:
        print(f"Error creating notification: {e}")
        return None


def create_admin_notification(title, message, notification_type, submission=None):
    """Create notification for all admin users"""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    try:
        # Get all admin users
        admin_users = User.objects.filter(
            Q(is_superuser=True) | Q(groups__name='Admin')
        ).distinct()
        
        # Create notification for each admin
        notifications = []
        for admin in admin_users:
            notif = create_notification(
                user=admin,
                title=title,
                message=message,
                notification_type=notification_type,
                submission=submission
            )
            if notif:
                notifications.append(notif)
        
        return notifications
        
    except Exception as e:
        print(f"Error creating admin notifications: {e}")
        return []

#----END OF NOTIFICATIONS HELPERS----



#----CATEGORIZATION----

@login_required
def folder_view(request):
    """Main folder view showing all categories"""
    categories = FileCategory.objects.all()
    
    # Calculate file counts for each category
    for category in categories:
        category.file_count = CategorizedFile.objects.filter(
            category=category,
            is_archived=False
        ).count()
    
    context = {
        'categories': categories,
    }
    return render(request, 'folder.html', context)


@login_required
def certification_files_view(request):
    """View for certification files"""
    # Get certificate-related categories
    categories = FileCategory.objects.filter(
        name__in=['certificates', 'ids', 'signatures']
    )
    
    for category in categories:
        category.file_count = CategorizedFile.objects.filter(
            category=category,
            is_archived=False
        ).count()
    
    context = {
        'categories': categories,
        'page_title': 'Certification Files',
    }
    return render(request, 'certification_filess.html', context)


@login_required
def monitoring_files_view(request):
    """View for monitoring/requirements files"""
    barangays = Barangay.objects.all()
    
    # Get requirement-related categories
    categories = FileCategory.objects.filter(
        name__in=['weekly', 'monthly', 'quarterly', 'semestral', 'annually']
    )
    
    for category in categories:
        category.file_count = CategorizedFile.objects.filter(
            category=category,
            is_archived=False
        ).count()
    
    context = {
        'barangays': barangays,
        'categories': categories,
        'page_title': 'Monitoring Files',
    }
    return render(request, 'monitoring_files.html', context)


@login_required
@require_http_methods(["GET"])
def api_category_files(request, category_name):
    """API endpoint to get files by category"""
    try:
        category = get_object_or_404(FileCategory, name=category_name)
        
        # Get filter parameters
        barangay_id = request.GET.get('barangay_id')
        search_query = request.GET.get('search', '').strip()
        file_type = request.GET.get('file_type', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        
        # Base query
        files = CategorizedFile.objects.filter(
            category=category,
            is_archived=False
        ).select_related('barangay', 'uploaded_by')
        
        # Apply filters
        if barangay_id:
            files = files.filter(barangay_id=barangay_id)
        
        if search_query:
            files = files.filter(
                Q(original_filename__icontains=search_query) |
                Q(detected_content__icontains=search_query) |
                Q(tags__icontains=search_query)
            )
        
        if file_type:
            files = files.filter(file_type=file_type)
        
        if date_from:
            files = files.filter(uploaded_at__date__gte=date_from)
        
        if date_to:
            files = files.filter(uploaded_at__date__lte=date_to)
        
        # Paginate
        paginator = Paginator(files, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        # Prepare response data
        files_data = []
        for file in page_obj:
            files_data.append({
                'id': file.id,
                'filename': file.original_filename,
                'file_url': file.file.url,
                'file_type': file.file_type,
                'file_size': file.file_size_mb,
                'detected_content': file.detected_content,
                'barangay': file.barangay.name if file.barangay else None,
                'period': file.period,
                'uploaded_at': file.uploaded_at.strftime('%B %d, %Y %I:%M %p'),
                'uploaded_by': file.uploaded_by.get_full_name() if file.uploaded_by else 'System',
                'tags': file.tags,
                'thumbnail': file.get_thumbnail_url(),
            })
        
        return JsonResponse({
            'success': True,
            'files': files_data,
            'total_count': paginator.count,
            'page': page_obj.number,
            'total_pages': paginator.num_pages,
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_upload_file(request):
    """API endpoint for manual file upload"""
    try:
        category_name = request.POST.get('category')
        barangay_id = request.POST.get('barangay_id')
        period = request.POST.get('period', '')
        tags = request.POST.get('tags', '')
        
        if not category_name:
            return JsonResponse({
                'success': False,
                'error': 'Category is required'
            }, status=400)
        
        if 'file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'error': 'No file uploaded'
            }, status=400)
        
        uploaded_file = request.FILES['file']
        
        # Get category
        category = get_object_or_404(FileCategory, name=category_name)
        
        # Determine file type
        mime_type = uploaded_file.content_type
        if mime_type.startswith('image/'):
            file_type = 'image'
        elif mime_type == 'application/pdf':
            file_type = 'pdf'
        elif mime_type.startswith('application/'):
            file_type = 'document'
        else:
            file_type = 'other'
        
        # Create categorized file
        categorized_file = CategorizedFile.objects.create(
            file=uploaded_file,
            original_filename=uploaded_file.name,
            file_type=file_type,
            file_size=uploaded_file.size,
            mime_type=mime_type,
            category=category,
            source='manual',
            barangay_id=barangay_id if barangay_id else None,
            period=period,
            uploaded_by=request.user,
            tags=tags,
        )
        
        # Update category file count
        category.update_file_count()
        
        # Log the upload
        try:
            AuditLog.objects.create(
                user=request.user,
                action='CREATE',
                content_object=categorized_file,
                description=f"Uploaded file to {category.display_name}: {uploaded_file.name}"
            )
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': 'File uploaded successfully',
            'file': {
                'id': categorized_file.id,
                'filename': categorized_file.original_filename,
                'file_url': categorized_file.file.url,
                'file_size': categorized_file.file_size_mb,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@require_http_methods(["DELETE"])
def api_delete_file(request, file_id):
    """
    FIXED: API endpoint to delete a file
    """
    try:
        print(f"\n{'='*50}")
        print(f"DELETE FILE - ID: {file_id}")
        print(f"{'='*50}")
        
        file = MonitoringFile.objects.get(id=file_id)
        filename = file.filename
        print(f"Found file: {filename}")
        
        # Delete the actual file
        try:
            if file.file:
                file.file.delete(save=False)
                print(f"✓ Deleted physical file")
        except Exception as file_del_error:
            print(f"✗ Could not delete physical file: {file_del_error}")
        
        # Delete database record
        file.delete()
        print(f"✓ Deleted database record")
        print(f"{'='*50}\n")
        
        return JsonResponse({
            'success': True,
            'message': f'File "{filename}" deleted successfully'
        })
        
    except MonitoringFile.DoesNotExist:
        print(f"✗ File {file_id} not found in database")
        return JsonResponse({
            'success': False,
            'error': 'File not found'
        }, status=404)
        
    except Exception as e:
        print(f"\n{'='*50}")
        print(f"ERROR IN DELETE FILE")
        print(f"{'='*50}")
        print(f"Error: {str(e)}")
        print(traceback.format_exc())
        print(f"{'='*50}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_archive_file(request, file_id):
    """API endpoint to archive a file"""
    try:
        file = get_object_or_404(CategorizedFile, id=file_id)
        
        file.archive()
        
        # Update category count
        file.category.update_file_count()
        
        return JsonResponse({
            'success': True,
            'message': 'File archived successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_move_file(request, file_id):
    """API endpoint to move file to different category"""
    try:
        file = get_object_or_404(CategorizedFile, id=file_id)
        
        data = json.loads(request.body)
        new_category_name = data.get('category')
        
        if not new_category_name:
            return JsonResponse({
                'success': False,
                'error': 'Category is required'
            }, status=400)
        
        new_category = get_object_or_404(FileCategory, name=new_category_name)
        old_category = file.category
        
        file.category = new_category
        file.save()
        
        # Update both category counts
        old_category.update_file_count()
        new_category.update_file_count()
        
        return JsonResponse({
            'success': True,
            'message': f'File moved to {new_category.display_name}'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def api_file_statistics(request):
    """API endpoint to get file statistics"""
    try:
        stats = {
            'total_files': CategorizedFile.objects.filter(is_archived=False).count(),
            'total_size_mb': round(
                CategorizedFile.objects.filter(is_archived=False).aggregate(
                    total=Sum('file_size')
                )['total'] / (1024 * 1024), 2
            ) if CategorizedFile.objects.exists() else 0,
            'by_category': [],
            'by_type': [],
            'recent_uploads': [],
        }
        
        # Stats by category
        for category in FileCategory.objects.all():
            count = CategorizedFile.objects.filter(
                category=category,
                is_archived=False
            ).count()
            if count > 0:
                stats['by_category'].append({
                    'name': category.display_name,
                    'count': count
                })
        
        # Stats by file type
        type_counts = CategorizedFile.objects.filter(
            is_archived=False
        ).values('file_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        stats['by_type'] = list(type_counts)
        
        # Recent uploads
        recent = CategorizedFile.objects.filter(
            is_archived=False
        ).order_by('-uploaded_at')[:5]
        
        for file in recent:
            stats['recent_uploads'].append({
                'filename': file.original_filename,
                'category': file.category.display_name,
                'uploaded_at': file.uploaded_at.strftime('%B %d, %Y'),
            })
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)




@require_http_methods(["GET"])
def get_files_by_category(request, category):
    """
    FIXED: API endpoint to get files by category
    URL: /api/files/category/<category>/
    """
    try:
        print(f"\n{'='*50}")
        print(f"GET FILES BY CATEGORY - START")
        print(f"{'='*50}")
        print(f"Category: {category}")
        print(f"Request method: {request.method}")
        print(f"GET params: {dict(request.GET)}")
        
        # Get query parameters
        barangay_id = request.GET.get('barangay_id', None)
        date_from = request.GET.get('date_from', None)
        date_to = request.GET.get('date_to', None)
        
        # Check if MonitoringFile model exists
        try:
            # Test query to check if table exists
            test_count = MonitoringFile.objects.count()
            print(f"✓ MonitoringFile table exists. Total records: {test_count}")
        except Exception as model_error:
            print(f"✗ MonitoringFile table error: {model_error}")
            # Return empty result gracefully
            return JsonResponse({
                'success': True,
                'files': [],
                'total_count': 0,
                'category': category,
                'message': 'No files table found'
            })
        
        # Base query
        files = MonitoringFile.objects.filter(category=category)
        print(f"Files in category '{category}': {files.count()}")
        
        # Apply filters
        if barangay_id:
            files = files.filter(barangay_id=barangay_id)
            print(f"After barangay filter ({barangay_id}): {files.count()}")
        
        if date_from:
            files = files.filter(uploaded_at__gte=date_from)
            print(f"After date_from filter: {files.count()}")
        
        if date_to:
            files = files.filter(uploaded_at__lte=date_to)
            print(f"After date_to filter: {files.count()}")
        
        # Order by most recent
        files = files.order_by('-uploaded_at')
        total_count = files.count()
        
        # Serialize files
        files_data = []
        for file_obj in files:
            try:
                # Build file data safely
                file_data = {
                    'id': file_obj.id,
                    'filename': getattr(file_obj, 'filename', 'Unknown'),
                    'file_url': '',
                    'file_type': 'unknown',
                    'file_size': 0,
                    'barangay': 'N/A',
                    'uploaded_at': ''
                }
                
                # Get file URL
                try:
                    if hasattr(file_obj, 'file') and file_obj.file:
                        file_data['file_url'] = file_obj.file.url
                        # Get file size
                        try:
                            file_data['file_size'] = round(file_obj.file.size / (1024 * 1024), 2)
                        except:
                            pass
                except:
                    pass
                
                # Get file type
                if file_obj.filename:
                    ext = file_obj.filename.lower().split('.')[-1]
                    if ext in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'svg', 'webp']:
                        file_data['file_type'] = 'image'
                    else:
                        file_data['file_type'] = 'document'
                
                # Get barangay name
                try:
                    if hasattr(file_obj, 'barangay') and file_obj.barangay:
                        file_data['barangay'] = file_obj.barangay.name
                except:
                    pass
                
                # Get upload date
                try:
                    if hasattr(file_obj, 'uploaded_at') and file_obj.uploaded_at:
                        file_data['uploaded_at'] = file_obj.uploaded_at.strftime('%Y-%m-%d %H:%M')
                except:
                    pass
                
                files_data.append(file_data)
                
            except Exception as file_error:
                print(f"✗ Error processing file {file_obj.id}: {file_error}")
                continue
        
        print(f"✓ Successfully processed {len(files_data)} files")
        print(f"{'='*50}\n")
        
        return JsonResponse({
            'success': True,
            'files': files_data,
            'total_count': total_count,
            'category': category
        })
        
    except Exception as e:
        print(f"\n{'='*50}")
        print(f"ERROR IN GET_FILES_BY_CATEGORY")
        print(f"{'='*50}")
        print(f"Error: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        print(f"Traceback:")
        print(traceback.format_exc())
        print(f"{'='*50}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e),
            'category': category
        }, status=500)


def get_file_type(filename):
    """Helper function to determine file type"""
    try:
        ext = filename.lower().split('.')[-1]
        image_extensions = ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'svg', 'webp']
        
        if ext in image_extensions:
            return 'image'
        return 'document'
    except:
        return 'unknown'

def test_monitoring_api(request):
    """
    Temporary test endpoint to verify setup
    URL: /test-monitoring/
    """
    try:
        # Test database connection
        from .models import MonitoringFile, Barangay
        
        monitoring_count = MonitoringFile.objects.count()
        barangay_count = Barangay.objects.count()
        
        # Get sample records
        sample_files = list(MonitoringFile.objects.values(
            'id', 'filename', 'category', 'barangay__name'
        )[:5])
        
        categories = MonitoringFile.objects.values_list('category', flat=True).distinct()
        
        return JsonResponse({
            'status': 'success',
            'database': 'connected',
            'monitoring_files_count': monitoring_count,
            'barangays_count': barangay_count,
            'categories': list(categories),
            'sample_files': sample_files,
        }, json_dumps_params={'indent': 2})
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500, json_dumps_params={'indent': 2})
    
@require_http_methods(["GET"])
def get_files_by_category_simple(request, category):
    """
    Get files by category (weekly, monthly, quarterly, semestral, annually)
    Files are automatically categorized when uploaded via requirements_monitoring
    """
    try:
        print(f"\n{'='*60}")
        print(f"📁 GET FILES BY CATEGORY: {category}")
        print(f"{'='*60}")
        
        # Get query parameters
        barangay_id = request.GET.get('barangay_id')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        page = int(request.GET.get('page', 1))
        
        print(f"Filters: barangay_id={barangay_id}, date_from={date_from}, date_to={date_to}")
        
        # Query CategorizedFile using period field
        # REMOVED is_archived filter since it doesn't exist
        files = CategorizedFile.objects.filter(
            period=category  # weekly, monthly, quarterly, semestral, annually
        ).select_related('barangay', 'uploaded_by', 'requirement_submission')
        
        print(f"Files with period '{category}': {files.count()}")
        
        # Apply filters
        if barangay_id:
            files = files.filter(barangay_id=barangay_id)
            print(f"After barangay filter: {files.count()}")
        
        if date_from:
            files = files.filter(uploaded_at__gte=date_from)
            print(f"After date_from filter: {files.count()}")
        
        if date_to:
            files = files.filter(uploaded_at__lte=date_to)
            print(f"After date_to filter: {files.count()}")
        
        # Order by newest first
        files = files.order_by('-uploaded_at')
        total_count = files.count()
        
        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(files, 20)
        page_obj = paginator.get_page(page)
        
        print(f"Total count: {total_count}, Page: {page}/{paginator.num_pages}")
        
        # Build response
        files_data = []
        for file_obj in page_obj:
            try:
                files_data.append({
                    'id': file_obj.id,
                    'filename': file_obj.original_filename,
                    'file_url': file_obj.file.url if file_obj.file else '',
                    'file_type': file_obj.file_type,
                    'file_size': file_obj.file_size_mb,
                    'barangay': file_obj.barangay.name if file_obj.barangay else 'N/A',
                    'uploaded_at': file_obj.uploaded_at.strftime('%B %d, %Y %I:%M %p'),
                    'uploaded_by': file_obj.uploaded_by.get_full_name() if file_obj.uploaded_by else 'System',
                    'detected_content': file_obj.detected_content or 'N/A',
                    'requirement_title': (
                        file_obj.requirement_submission.requirement.title 
                        if file_obj.requirement_submission 
                        else 'N/A'
                    ),
                    'tags': file_obj.tags or '',
                })
            except Exception as file_err:
                print(f"✗ Error processing file {file_obj.id}: {file_err}")
                continue
        
        print(f"✓ Successfully processed {len(files_data)} files")
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': True,
            'files': files_data,
            'total_count': total_count,
            'page': page,
            'total_pages': paginator.num_pages,
            'category': category
        })
        
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"❌ ERROR IN get_files_by_category_simple")
        print(f"{'='*60}")
        print(f"Error: {str(e)}")
        import traceback
        print(traceback.format_exc())
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e),
            'category': category
        }, status=500)
    

@require_http_methods(["DELETE"])  
def api_delete_monitoring_file(request, file_id):
    """
    Delete a categorized file
    Works for files uploaded via requirements_monitoring
    """
    try:
        print(f"\n{'='*60}")
        print(f"🗑️ DELETE FILE REQUEST")
        print(f"{'='*60}")
        print(f"File ID: {file_id}")
        print(f"User: {request.user}")
        
        # Get the file from CategorizedFile
        file = CategorizedFile.objects.get(id=file_id)
        filename = file.original_filename
        category = file.category.display_name if file.category else 'Unknown'
        
        print(f"Found file: {filename}")
        print(f"Category: {category}")
        
        # Delete physical file from storage
        try:
            if file.file:
                file.file.delete(save=False)
                print(f"✓ Deleted physical file from storage")
        except Exception as storage_err:
            print(f"⚠️ Could not delete physical file: {storage_err}")
        
        # Delete the RequirementAttachment if it exists
        if file.requirement_attachment:
            try:
                file.requirement_attachment.delete()
                print(f"✓ Deleted linked RequirementAttachment")
            except Exception as att_err:
                print(f"⚠️ Could not delete RequirementAttachment: {att_err}")
        
        # Delete database record
        file.delete()
        print(f"✓ Deleted from database: {filename}")
        
        # Update category file count
        if file.category:
            file.category.update_file_count()
            print(f"✓ Updated category file count")
        
        # Log the deletion
        try:
            AuditLog.objects.create(
                user=request.user,
                action='DELETE',
                description=f"Deleted file: {filename} from {category}"
            )
        except:
            pass
        
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': True,
            'message': f'File "{filename}" deleted successfully'
        })
        
    except CategorizedFile.DoesNotExist:
        print(f"✗ File {file_id} not found in CategorizedFile table")
        print(f"{'='*60}\n")
        return JsonResponse({
            'success': False,
            'error': 'File not found'
        }, status=404)
        
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"❌ ERROR IN api_delete_monitoring_file")
        print(f"{'='*60}")
        print(f"Error: {str(e)}")
        print(traceback.format_exc())
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

    



@require_http_methods(["GET"])
def get_category_file_counts(request):
    """
    Get file counts for all categories
    Useful for updating the monitoring_files.html folder counts
    """
    try:
        categories = ['weekly', 'monthly', 'quarterly', 'semestral', 'annually']
        barangay_id = request.GET.get('barangay_id')
        
        counts = {}
        for category in categories:
            query = CategorizedFile.objects.filter(
                period=category,
                is_archived=False
            )
            
            if barangay_id:
                query = query.filter(barangay_id=barangay_id)
            
            counts[category] = query.count()
        
        return JsonResponse({
            'success': True,
            'counts': counts
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)





